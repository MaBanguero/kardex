from django.db import transaction
from django.core.exceptions import ValidationError
from django.db.models import Sum
from .models import Documento, DocumentoDetalle, InventarioStock, Ubicacion, Medicamento, PerfilUsuario
from django.contrib.auth.models import User, Group
import io, csv, re
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side


def procesar_carga_masiva_usuarios(archivo_csv):
    """
    Formato CSV esperado: username, first_name, last_name, email, roles, identificacion, ubicacion_id, password
    roles debe ser separado por pipes: ADMIN|REGENTE|ENFERMERA
    """
    decoded_file = archivo_csv.read().decode('utf-8')
    io_string = io.StringIO(decoded_file)
    reader = csv.DictReader(io_string)

    contador = 0
    with transaction.atomic():
        for row in reader:
            if User.objects.filter(username=row['username']).exists():
                continue

            user = User.objects.create_user(
                username=row['username'],
                first_name=row['first_name'],
                last_name=row['last_name'],
                email=row['email'],
                password=row['password']
            )

            ubicacion = Ubicacion.objects.get(id=row['ubicacion_id'])

            PerfilUsuario.objects.create(
                usuario=user,
                ubicacion_asignada=ubicacion,
                numero_identificacion=row['identificacion']
            )

            # Asignar roles via grupos
            roles = row.get('roles', '').split('|')
            for rol_name in roles:
                rol_name = rol_name.strip().upper()
                if rol_name in ('ADMIN', 'REGENTE', 'ENFERMERA'):
                    grupo, _ = Group.objects.get_or_create(name=rol_name)
                    user.groups.add(grupo)

            contador += 1
    return contador

def procesar_traslado(usuario_id, ubicacion_origen_id, ubicacion_destino_id, lote, medicamento_id, cantidad_a_mover):
    try:
        with transaction.atomic():
            stock_origen = InventarioStock.objects.select_for_update().get(
                ubicacion_id=ubicacion_origen_id, medicamento_id=medicamento_id, lote=lote
            )
            if stock_origen.cantidad_actual < cantidad_a_mover:
                raise ValidationError("Stock insuficiente.")
            stock_destino, created = InventarioStock.objects.select_for_update().get_or_create(
                ubicacion_id=ubicacion_destino_id, medicamento_id=medicamento_id, lote=lote,
                defaults={'cantidad_actual': 0, 'fecha_vencimiento': stock_origen.fecha_vencimiento}
            )
            stock_origen.cantidad_actual -= cantidad_a_mover
            stock_origen.save()
            stock_destino.cantidad_actual += cantidad_a_mover
            stock_destino.save()
            documento = Documento.objects.create(
                tipo_mov='TRASLADO', origen_id=ubicacion_origen_id, destino_id=ubicacion_destino_id, usuario_id=usuario_id
            )
            DocumentoDetalle.objects.create(documento=documento, medicamento_id=medicamento_id, lote=lote, cantidad=cantidad_a_mover)
        return documento
    except Exception as e:
        raise ValidationError(str(e))


# kardex/services.py

def registrar_devolucion(usuario, doc_salida_id, cantidad_devolver):
    """
    Procesa la devolución parcial o total de medicamentos, reingresándolos al stock
    y vinculándolos con la salida original para auditoría.
    """
    try:
        with transaction.atomic():
            # Bloqueamos la salida original para evitar registros duplicados simultáneos
            salida_original = Documento.objects.select_for_update().get(id=doc_salida_id, tipo_mov='SALIDA')

            # Validación de tiempo (regla de negocio)
            if salida_original.tiempo_agotado_para_devolucion():
                raise ValidationError("El tiempo límite para devoluciones ha expirado.")

            # Obtenemos el detalle del medicamento que se retiró
            detalle_original = salida_original.detalles.first()
            if not detalle_original:
                raise ValidationError("No se encontró el registro de medicamento original.")

            # Validación: No se puede devolver más de lo que salió originalmente
            total_ya_devuelto = DocumentoDetalle.objects.filter(
                documento__documento_referencia=salida_original,
                medicamento=detalle_original.medicamento,
                lote=detalle_original.lote
            ).aggregate(total=Sum('cantidad'))['total'] or 0

            if (total_ya_devuelto + cantidad_devolver) > detalle_original.cantidad:
                raise ValidationError(
                    f"Error: No puedes devolver {cantidad_devolver} unidades. Solo se retiraron {detalle_original.cantidad} y ya has devuelto {total_ya_devuelto}.")

            # 1. Crear el documento de DEVOLUCIÓN
            doc_devolucion = Documento.objects.create(
                tipo_mov='DEVOLUCION',
                usuario=usuario,
                destino=salida_original.origen,  # Regresa a la ubicación de donde salió
                id_paciente=salida_original.id_paciente,
                documento_referencia=salida_original
            )

            # 2. Registrar el detalle del reingreso
            DocumentoDetalle.objects.create(
                documento=doc_devolucion,
                medicamento=detalle_original.medicamento,
                lote=detalle_original.lote,
                cantidad=cantidad_devolver
            )

            # 3. Actualizar el Stock físicamente
            stock, _ = InventarioStock.objects.select_for_update().get_or_create(
                ubicacion=salida_original.origen,
                medicamento=detalle_original.medicamento,
                lote=detalle_original.lote
            )
            stock.cantidad_actual += cantidad_devolver
            stock.save()

            return doc_devolucion

    except Documento.DoesNotExist:
        raise ValidationError("No se encontró el documento de salida original.")


def registrar_devolucion_agrupada(usuario, nombre_medicamento, cantidad_a_devolver, id_paciente):
    """
    Busca todas las salidas pendientes del paciente y distribuye la devolución
    creando un enlace (documento_referencia) por cada documento afectado.
    """
    ubicacion_id = usuario.perfil.ubicacion_asignada.id

    try:
        with transaction.atomic():
            # 1. Buscar todos los detalles de salida de este paciente para este medicamento
            # Ordenados de la salida más reciente a la más antigua (LIFO para devoluciones)
            detalles_salida = DocumentoDetalle.objects.select_for_update().filter(
                documento__usuario=usuario,
                documento__tipo_mov='SALIDA',
                documento__id_paciente=id_paciente,
                medicamento__principio_activo__iexact=nombre_medicamento.strip()
            ).order_by('-documento__fecha')

            # 2. Calcular saldos globales para validar
            total_salidas = detalles_salida.aggregate(total=Sum('cantidad'))['total'] or 0

            total_ya_devuelto = DocumentoDetalle.objects.filter(
                documento__documento_referencia__in=detalles_salida.values_list('documento', flat=True),
                documento__tipo_mov='DEVOLUCION',
                medicamento__principio_activo__iexact=nombre_medicamento.strip()
            ).aggregate(total=Sum('cantidad'))['total'] or 0

            saldo_disponible_global = total_salidas - total_ya_devuelto

            if cantidad_a_devolver > saldo_disponible_global:
                raise ValidationError(
                    f"Error: No puedes devolver {cantidad_a_devolver}. El saldo pendiente del paciente es de {saldo_disponible_global} unidades.")

            cantidad_pendiente = cantidad_a_devolver

            # 3. Distribuir la devolución creando documentos individuales para mantener el historial
            for detalle in detalles_salida:
                if cantidad_pendiente <= 0:
                    break  # Ya terminamos de devolver todo lo solicitado

                # Calcular cuánto se ha devuelto específicamente de ESTE documento de salida
                ya_devuelto_este_detalle = DocumentoDetalle.objects.filter(
                    documento__documento_referencia=detalle.documento,
                    medicamento=detalle.medicamento,
                    lote=detalle.lote
                ).aggregate(total=Sum('cantidad'))['total'] or 0

                disponible_en_este_doc = detalle.cantidad - ya_devuelto_este_detalle

                if disponible_en_este_doc <= 0:
                    continue  # Esta salida ya fue devuelta en su totalidad, pasamos a la siguiente

                # Tomamos lo que podamos de esta salida
                cantidad_tramo = min(cantidad_pendiente, disponible_en_este_doc)

                # ¡CLAVE!: Creamos el documento de devolución apuntando a la salida específica
                doc_dev = Documento.objects.create(
                    tipo_mov='DEVOLUCION',
                    usuario=usuario,
                    destino_id=ubicacion_id,
                    id_paciente=id_paciente,
                    documento_referencia=detalle.documento  # <- ESTO HACE QUE APAREZCA LA NOTA VERDE EN EL HISTORIAL
                )

                # Creamos el detalle específico
                DocumentoDetalle.objects.create(
                    documento=doc_dev,
                    medicamento=detalle.medicamento,
                    lote=detalle.lote,
                    cantidad=cantidad_tramo
                )

                # Devolvemos el stock físico
                stock = InventarioStock.objects.get(
                    ubicacion_id=ubicacion_id,
                    medicamento=detalle.medicamento,
                    lote=detalle.lote
                )
                stock.cantidad_actual += cantidad_tramo
                stock.save()

                # Restamos a la meta de devolución
                cantidad_pendiente -= cantidad_tramo

            return True  # Finalizó con éxito

    except Exception as e:
        raise ValidationError(str(e))


def registrar_salida_paciente(usuario, stock_id, cantidad, id_paciente):
    """
    Resta stock de la ubicación y crea el documento de salida.
    """
    with transaction.atomic():
        # Bloqueamos la fila para evitar que otra enfermera use el mismo stock
        stock = InventarioStock.objects.select_for_update().get(id=stock_id)

        if stock.cantidad_actual < cantidad:
            raise ValidationError(f"Stock insuficiente. Solo hay {stock.cantidad_actual} unidades.")

        # 1. Descontar del inventario
        stock.cantidad_actual -= cantidad
        stock.save()

        # 2. Crear la cabecera del documento
        doc = Documento.objects.create(
            tipo_mov='SALIDA',
            usuario=usuario,
            origen=stock.ubicacion,
            id_paciente=id_paciente
        )

        # 3. Crear el detalle del movimiento
        DocumentoDetalle.objects.create(
            documento=doc,
            medicamento=stock.medicamento,
            lote=stock.lote,
            cantidad=cantidad
        )
        return doc

def registrar_salida_paciente_inteligente(usuario, nombre_medicamento, cantidad_solicitada, id_paciente, cups_codigo=None):
    """
    Descuenta stock automáticamente del lote más próximo a vencer (FEFO).
    Si se proporciona cups_codigo, busca por CUPS; de lo contrario busca por nombre.
    """
    ubicacion_id = usuario.perfil.ubicacion_asignada.id

    try:
        with transaction.atomic():
            # 1. Buscamos todos los lotes de este medicamento en esta sede
            # Ordenados por fecha_vencimiento (El más viejo primero)
            filtro = {
                'ubicacion_id': ubicacion_id,
                'cantidad_actual__gt': 0,
            }
            if cups_codigo:
                # Buscar por cups_codigo; si no hay match, probar por codigo (CUM)
                stocks_disponibles = InventarioStock.objects.select_for_update().filter(
                    ubicacion_id=ubicacion_id,
                    medicamento__cups_codigo=cups_codigo,
                    cantidad_actual__gt=0
                ).order_by('fecha_vencimiento')

                if not stocks_disponibles.exists():
                    # Fallback: buscar por codigo (CUM) ya que el frontend
                    # usa codigo como cups_codigo cuando no hay CUPS explícito
                    stocks_disponibles = InventarioStock.objects.select_for_update().filter(
                        ubicacion_id=ubicacion_id,
                        medicamento__codigo=cups_codigo,
                        cantidad_actual__gt=0
                    ).order_by('fecha_vencimiento')
            else:
                filtro['medicamento__principio_activo__iexact'] = nombre_medicamento.strip()

                stocks_disponibles = InventarioStock.objects.select_for_update().filter(
                    **filtro
                ).order_by('fecha_vencimiento')

            # 2. Verificamos si la suma de todos los lotes alcanza
            total_disponible = sum(stock.cantidad_actual for stock in stocks_disponibles)
            if total_disponible < cantidad_solicitada:
                raise ValidationError(f"Stock insuficiente. Solo hay {total_disponible} unidades disponibles en total.")

            # 3. Creamos la cabecera del documento
            doc = Documento.objects.create(
                tipo_mov='SALIDA',
                usuario=usuario,
                origen_id=ubicacion_id,
                id_paciente=id_paciente
            )

            # 4. Descontamos lote por lote hasta cumplir la cuota (FEFO)
            cantidad_restante = cantidad_solicitada

            for stock in stocks_disponibles:
                if cantidad_restante <= 0:
                    break # Ya cumplimos con lo solicitado

                # Calculamos cuánto podemos sacarle a este lote
                cantidad_a_descontar = min(stock.cantidad_actual, cantidad_restante)

                # Restamos
                stock.cantidad_actual -= cantidad_a_descontar
                stock.save()

                # Creamos el detalle específico con este lote
                DocumentoDetalle.objects.create(
                    documento=doc,
                    medicamento=stock.medicamento,
                    lote=stock.lote,
                    cantidad=cantidad_a_descontar
                )

                # Actualizamos la meta
                cantidad_restante -= cantidad_a_descontar

            return doc

    except Exception as e:
        raise ValidationError(str(e))


def calcular_semaforo(fecha_vencimiento):
    """
    Calcula el color de semáforo INVIMA según la Resolución 3100 de 2019.
    Retorna VERDE | AMARILLO | NARANJA | ROJO | VENCIDO.
    """
    if not fecha_vencimiento:
        return ''
    if hasattr(fecha_vencimiento, 'date'):
        fecha_venc = fecha_vencimiento.date()
    else:
        fecha_venc = fecha_vencimiento

    dias_restantes = (fecha_venc - datetime.date.today()).days

    if dias_restantes < 0:
        return 'VENCIDO'
    elif dias_restantes < 30:
        return 'ROJO'
    elif dias_restantes < 90:
        return 'NARANJA'
    elif dias_restantes < 180:
        return 'AMARILLO'
    else:
        return 'VERDE'


def generar_excel_kardex(mes, anio, ubicacion_id, tipo='MEDICAMENTO'):
    if tipo == 'DISPOSITIVO':
        return generar_excel_dispositivos(mes, anio, ubicacion_id)
    return generar_excel_medicamentos(mes, anio, ubicacion_id)


def _write_medicamentos_kardex(ws, mes, anio, ubicacion_id, hoy=None):
    """Escribe contenido del kardex de medicamentos en un worksheet existente."""
    if hoy is None:
        hoy = datetime.datetime.now()
    ubicacion = Ubicacion.objects.get(id=ubicacion_id)
    unidad_nombre = ubicacion.nombre

    fuente_negrita = Font(bold=True, size=9, name='Arial')
    fuente_normal = Font(size=9, name='Arial')
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alineacion_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
    fondo_gris = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    def estilizar_rango(rango, fuente, alineacion, fondo, borde):
        for fila in ws[rango]:
            for celda in fila:
                if fuente: celda.font = fuente
                if alineacion: celda.alignment = alineacion
                if fondo: celda.fill = fondo
                if borde: celda.border = borde

    stocks = InventarioStock.objects.filter(
        ubicacion_id=ubicacion_id,
        medicamento__tipo='MEDICAMENTO'
    ).select_related('medicamento')

    datos_kardex = []
    max_ingresos = 1
    max_egresos = 1

    for stock in stocks:
        ingresos_qs = DocumentoDetalle.objects.filter(
            documento__tipo_mov__in=['ENTRADA', 'DEVOLUCION'],
            documento__destino_id=ubicacion_id,
            medicamento=stock.medicamento,
            lote=stock.lote,
            documento__fecha__month=mes,
            documento__fecha__year=anio
        ).order_by('documento__fecha')

        egresos_qs = DocumentoDetalle.objects.filter(
            documento__tipo_mov__in=['SALIDA', 'TRASLADO'],
            documento__origen_id=ubicacion_id,
            medicamento=stock.medicamento,
            lote=stock.lote,
            documento__fecha__month=mes,
            documento__fecha__year=anio
        ).order_by('documento__fecha')

        ingresos_list = list(ingresos_qs)
        egresos_list = list(egresos_qs)

        if len(ingresos_list) > max_ingresos: max_ingresos = len(ingresos_list)
        if len(egresos_list) > max_egresos: max_egresos = len(egresos_list)

        total_ingresos = sum(mov.cantidad for mov in ingresos_list)
        total_egresos = sum(mov.cantidad for mov in egresos_list)

        saldo_final = stock.cantidad_actual
        saldo_inicial = saldo_final - total_ingresos + total_egresos

        datos_kardex.append({
            'stock': stock,
            'ingresos': ingresos_list,
            'egresos': egresos_list,
            'total_ingresos': total_ingresos,
            'total_egresos': total_egresos,
            'saldo_inicial': saldo_inicial,
            'saldo_final': saldo_final
        })

    col_saldo_ini = 12
    col_ingresos_start = 13
    col_ingresos_end = col_ingresos_start + (max_ingresos * 2) - 1
    col_total_ingresos = col_ingresos_end + 1
    col_egresos_start = col_total_ingresos + 1
    col_egresos_end = col_egresos_start + (max_egresos * 2) - 1
    col_total_egresos = col_egresos_end + 1
    col_saldo_final = col_total_egresos + 1
    col_verificacion = col_saldo_final + 1

    meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE",
             "NOVIEMBRE", "DICIEMBRE"]
    ws['K1'] = "Código:"; ws['L1'] = "PM-SF-FR12"
    ws['K2'] = "Versión:"; ws['L2'] = "1"
    ws['K3'] = "Fecha Act.:"; ws['L3'] = hoy.strftime("%d.%m.%Y")
    for row in range(1, 4):
        ws[f'K{row}'].font = fuente_negrita
        ws[f'K{row}'].alignment = Alignment(horizontal="right")
        ws[f'L{row}'].font = fuente_normal

    ws.merge_cells('D1:H3')
    ws['D1'] = "KARDEX DE MEDICAMENTOS"
    ws['D1'].font = Font(bold=True, size=14, name='Arial')
    ws['D1'].alignment = alineacion_centro

    ws['A5'] = "FECHA:"; ws['B5'] = f"{hoy.day} DE {meses[hoy.month - 1]} DE {hoy.year}"
    ws['D5'] = "ÁREA/SERVICIO:"; ws['E5'] = "FARMACIA"
    ws['G5'] = "UBICACIÓN:"; ws['H5'] = unidad_nombre
    ws['J5'] = "UNIDAD DE ATENCIÓN:"; ws['K5'] = "PUERTO TEJADA"
    for c in ['A5', 'D5', 'G5', 'J5']:
        ws[c].font = fuente_negrita; ws[c].alignment = Alignment(horizontal="right", vertical="center")
    for c in ['B5', 'E5', 'H5', 'K5']:
        ws[c].font = fuente_normal

    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 40

    columnas_base = [
        ("A", "ITEM"), ("B", "MEDICAMENTOS (PRINCIPIO ACTIVO)"), ("C", "FORMA FARMACÉUTICA"),
        ("D", "CONCENTRACIÓN"), ("E", "PRESENTACIÓN COMERCIAL"),
        ("F", "LOTE"), ("G", "FECHA DE VENCIMIENTO"),
        ("H", "UNIDAD DE MEDIDA"), ("I", "REGISTRO INVIMA"),
        ("J", "SEMAFORIZACIÓN"), ("K", "MEDICAMENTO LASA")
    ]
    for letra, titulo in columnas_base:
        rango = f"{letra}7:{letra}8"
        ws.merge_cells(rango)
        ws[f"{letra}7"] = titulo
        estilizar_rango(rango, fuente_negrita, alineacion_centro, fondo_gris, borde_fino)

    grupos = [
        (f"{get_column_letter(col_saldo_ini)}7:{get_column_letter(col_saldo_ini)}8",
         "SALDOS INICIO DEL PERIODO\nCANTIDAD"),
        (f"{get_column_letter(col_ingresos_start)}7:{get_column_letter(col_ingresos_end)}7", "INGRESOS"),
        (f"{get_column_letter(col_total_ingresos)}7:{get_column_letter(col_total_ingresos)}8", "TOTAL\nINGRESOS"),
        (f"{get_column_letter(col_egresos_start)}7:{get_column_letter(col_egresos_end)}7", "EGRESOS"),
        (f"{get_column_letter(col_total_egresos)}7:{get_column_letter(col_total_egresos)}8", "TOTAL\nEGRESOS"),
        (f"{get_column_letter(col_saldo_final)}7:{get_column_letter(col_saldo_final)}8", "SALDOS AL FINAL\nDEL PERIODO"),
        (f"{get_column_letter(col_verificacion)}7:{get_column_letter(col_verificacion)}8", "VERIFICACION EXISTENCIAS\nEN FISICO"),
    ]
    for rango, texto in grupos:
        ws.merge_cells(rango); ws[rango.split(':')[0]] = texto
        estilizar_rango(rango, fuente_negrita, alineacion_centro, fondo_gris, borde_fino)

    for col in range(col_ingresos_start, col_ingresos_end + 1, 2):
        ws.cell(row=8, column=col, value="FECHA"); ws.cell(row=8, column=col + 1, value="CANT.")
    for col in range(col_egresos_start, col_egresos_end + 1, 2):
        ws.cell(row=8, column=col, value="FECHA"); ws.cell(row=8, column=col + 1, value="CANT.")
    estilizar_rango(f"{get_column_letter(col_ingresos_start)}8:{get_column_letter(col_ingresos_end)}8", fuente_negrita,
                    alineacion_centro, fondo_gris, borde_fino)
    estilizar_rango(f"{get_column_letter(col_egresos_start)}8:{get_column_letter(col_egresos_end)}8", fuente_negrita,
                    alineacion_centro, fondo_gris, borde_fino)

    for letra, ancho in {'A': 5, 'B': 30, 'C': 18, 'D': 15, 'E': 12, 'F': 12, 'G': 10, 'H': 15, 'I': 10, 'J': 12, 'K': 12}.items():
        ws.column_dimensions[letra].width = ancho
    ws.column_dimensions[get_column_letter(col_saldo_ini)].width = 12
    ws.column_dimensions[get_column_letter(col_total_ingresos)].width = 11
    ws.column_dimensions[get_column_letter(col_total_egresos)].width = 11
    ws.column_dimensions[get_column_letter(col_saldo_final)].width = 12
    ws.column_dimensions[get_column_letter(col_verificacion)].width = 14

    fila_actual = 9
    for idx, data in enumerate(datos_kardex, start=1):
        ws.row_dimensions[fila_actual].height = 20
        m = data['stock'].medicamento; stock = data['stock']

        semaforo = calcular_semaforo(stock.fecha_vencimiento)

        # Columnas: ITEM, PRINCIPIO ACTIVO, FORMA FARM., CONCENTRACIÓN,
        # PRESENTACIÓN COMERCIAL, LOTE, FECHA VTO, UNIDAD MEDIDA,
        # REG INVIMA, SEMAFORIZACIÓN, LASA
        datos_fila = [
            idx,
            m.principio_activo,
            m.forma_farmaceutica,
            getattr(m, 'concentracion', ''),
            getattr(m, 'presentacion', ''),
            stock.lote,
            stock.fecha_vencimiento.strftime("%d/%m/%Y"),
            getattr(m, 'unidad_medida', ''),  # UNIDAD DE MEDIDA
            m.registro_invima or '',
            semaforo,  # SEMAFORIZACIÓN
            '',  # MEDICAMENTO LASA
        ]
        for col_idx, valor in enumerate(datos_fila, start=1):
            celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
            celda.font = fuente_normal
            celda.alignment = alineacion_centro if col_idx not in [2, 3] else alineacion_izq

        # Colorear celda de semáforo
        colores_semaforo = {
            'VENCIDO': 'FF0000', 'ROJO': 'FF4444',
            'NARANJA': 'FF8800', 'AMARILLO': 'FFDD00',
            'VERDE': '44BB44'
        }
        color_hex = colores_semaforo.get(semaforo, '')
        if color_hex:
            ws.cell(row=fila_actual, column=10).fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type='solid'
            )
            ws.cell(row=fila_actual, column=10).font = Font(
                bold=True, size=9, name='Arial',
                color='FFFFFF' if semaforo in ('VENCIDO', 'ROJO', 'NARANJA', 'VERDE') else '000000'
            )

        ws.cell(row=fila_actual, column=col_saldo_ini, value=data['saldo_inicial']).alignment = alineacion_centro
        c_ing = col_ingresos_start
        for mov in data['ingresos']:
            ws.cell(row=fila_actual, column=c_ing, value=mov.documento.fecha.strftime("%d/%m")).alignment = alineacion_centro
            ws.cell(row=fila_actual, column=c_ing + 1, value=mov.cantidad).alignment = alineacion_centro
            c_ing += 2
        ws.cell(row=fila_actual, column=col_total_ingresos, value=data['total_ingresos']).alignment = alineacion_centro
        c_eg = col_egresos_start
        for mov in data['egresos']:
            ws.cell(row=fila_actual, column=c_eg, value=mov.documento.fecha.strftime("%d/%m")).alignment = alineacion_centro
            ws.cell(row=fila_actual, column=c_eg + 1, value=mov.cantidad).alignment = alineacion_centro
            c_eg += 2
        ws.cell(row=fila_actual, column=col_total_egresos, value=data['total_egresos']).alignment = alineacion_centro
        ws.cell(row=fila_actual, column=col_saldo_final, value=data['saldo_final']).font = fuente_negrita
        ws.cell(row=fila_actual, column=col_saldo_final).alignment = alineacion_centro
        estilizar_rango(f"A{fila_actual}:{get_column_letter(col_verificacion)}{fila_actual}", None, None, None, borde_fino)
        fila_actual += 1

    ws.freeze_panes = 'A9'


def generar_excel_medicamentos(mes, anio, ubicacion_id):
    """Formato PM-SF-FR12: KARDEX DE MEDICAMENTOS"""
    wb = Workbook()
    ws = wb.active
    ws.title = "KARDEX"
    _write_medicamentos_kardex(ws, mes, anio, ubicacion_id)
    return wb


def _write_dispositivos_kardex(ws, mes, anio, ubicacion_id, hoy=None):
    """Escribe contenido del kardex de dispositivos médicos en un worksheet existente."""
    if hoy is None:
        hoy = datetime.datetime.now()
    ubicacion = Ubicacion.objects.get(id=ubicacion_id)
    unidad_nombre = ubicacion.nombre

    # Estilos
    fuente_negrita = Font(bold=True, size=9, name='Arial')
    fuente_normal = Font(size=9, name='Arial')
    fuente_titulo = Font(bold=True, size=14, name='Arial')
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alineacion_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    alineacion_der = Alignment(horizontal="right", vertical="center", wrap_text=False)
    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
    fondo_gris = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    def estilizar_rango(rango, fuente, alineacion, fondo, borde):
        for fila in ws[rango]:
            for celda in fila:
                if fuente: celda.font = fuente
                if alineacion: celda.alignment = alineacion
                if fondo: celda.fill = fondo
                if borde: celda.border = borde

    meses_es = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

    # 1. Preparar Datos
    stocks = InventarioStock.objects.filter(
        ubicacion_id=ubicacion_id,
        medicamento__tipo='DISPOSITIVO'
    ).select_related('medicamento')

    datos_kardex = []
    max_ingresos = 1
    max_egresos = 1

    for stock in stocks:
        ingresos_qs = DocumentoDetalle.objects.filter(
            documento__tipo_mov__in=['ENTRADA', 'DEVOLUCION'],
            documento__destino_id=ubicacion_id,
            medicamento=stock.medicamento,
            lote=stock.lote,
            documento__fecha__month=mes,
            documento__fecha__year=anio
        ).order_by('documento__fecha')

        egresos_qs = DocumentoDetalle.objects.filter(
            documento__tipo_mov__in=['SALIDA', 'TRASLADO'],
            documento__origen_id=ubicacion_id,
            medicamento=stock.medicamento,
            lote=stock.lote,
            documento__fecha__month=mes,
            documento__fecha__year=anio
        ).order_by('documento__fecha')

        ingresos_list = list(ingresos_qs)
        egresos_list = list(egresos_qs)

        if len(ingresos_list) > max_ingresos: max_ingresos = len(ingresos_list)
        if len(egresos_list) > max_egresos: max_egresos = len(egresos_list)

        total_ingresos = sum(mov.cantidad for mov in ingresos_list)
        total_egresos = sum(mov.cantidad for mov in egresos_list)

        saldo_final = stock.cantidad_actual
        saldo_inicial = saldo_final - total_ingresos + total_egresos

        datos_kardex.append({
            'stock': stock,
            'ingresos': ingresos_list,
            'egresos': egresos_list,
            'total_ingresos': total_ingresos,
            'total_egresos': total_egresos,
            'saldo_inicial': saldo_inicial,
            'saldo_final': saldo_final
        })

    # --- Columnas dinámicas ---
    COL_ITEM = 1
    COL_DESCRIPCION = 2
    COL_MARCA = 3
    COL_SERIE = 4
    COL_PRESENTACION = 5
    COL_REG_INVIMA = 6
    COL_VIDA_UTIL = 7
    COL_LOTE = 8
    COL_FECHA_VENCE = 9
    COL_SEMAFORIZACION = 10
    COL_CLASIF_RIESGO = 11
    COL_SALDO_INI = 12
    COL_ING_START = 13
    COL_ING_END = COL_ING_START + (max_ingresos * 2) - 1
    COL_TOTAL_ING = COL_ING_END + 1
    COL_EGR_START = COL_TOTAL_ING + 1
    COL_EGR_END = COL_EGR_START + (max_egresos * 2) - 1
    COL_TOTAL_EGR = COL_EGR_END + 1
    COL_SALDO_FIN = COL_TOTAL_EGR + 1
    COL_VERIFICACION = COL_SALDO_FIN + 1
    ULTIMA_COL = COL_VERIFICACION

    # ============================================================
    # FILAS 1-3: CÓDIGO, VERSIÓN, FECHA ACTUALIZACIÓN
    # ============================================================
    ws.merge_cells(start_row=1, start_column=4, end_row=3, end_column=8)
    ws['D1'] = "KARDEX DE DISPOSITIVOS MEDICOS E INSUMOS"
    ws['D1'].font = fuente_titulo
    ws['D1'].alignment = alineacion_centro

    ws['AF1'] = 'Código:'; ws['AG1'] = 'PM-SF-FR11'
    ws['AF2'] = 'Versión:'; ws['AG2'] = '1'
    ws['AF3'] = 'Fecha de Actualización:'; ws['AG3'] = '12.05.2025'
    for r in [1, 2, 3]:
        ws.cell(row=r, column=32).font = fuente_negrita
        ws.cell(row=r, column=32).alignment = alineacion_der
        ws.cell(row=r, column=33).font = fuente_normal

    # ============================================================
    # FILA 4: FECHA, ÁREA, UBICACIÓN, UNIDAD
    # ============================================================
    ws['A4'] = 'FECHA:'
    ws['B4'] = f"{hoy.day} DE {meses_es[hoy.month - 1]} DE {hoy.year}"
    ws['E4'] = 'AREA Y/O SERVICIO:'
    ws['F4'] = 'FARMACIA'
    ws['J4'] = 'UBICACIÓN:'
    ws['K4'] = unidad_nombre.upper()
    ws['N4'] = 'UNIDAD DE ATENCIÓN EN SALUD:'
    ws['O4'] = 'PADILLA'
    for c in ['A4', 'E4', 'J4', 'N4']:
        ws[c].font = fuente_negrita
        ws[c].alignment = alineacion_der
    for c in ['B4', 'F4', 'K4', 'O4']:
        ws[c].font = fuente_normal

    # ============================================================
    # FILA 6-7: ENCABEZADOS DE TABLA (doble fila)
    # ============================================================
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 35

    columnas_base = [
        (1, "ITEM"),
        (2, "DESCRIPCIÓN DEL DISPOSITIVO MEDICO\n(Corresponde al nombre que aparece en el registro sanitario"),
        (3, "MARCA DEL DISPOSITIVO"),
        (4, "SERIE\n(Cuando aplique)"),
        (5, "PRESENTACIÓN COMERCIAL"),
        (6, "REGISTRO INVIMA O PERMISO\nDE COMERCIALIZACION"),
        (7, "VIDA UTIL\n(Tiempo que el dispositivo médico e insumo mantiene\nlas condiciones y propiedades físicas y químicas inalteradas)"),
        (8, "LOTE"),
        (9, "FECHA DE \nVENCIMIENTO"),
        (10, "SEMAFORIZACION"),
        (11, "CLASIFICACION\nDE RIESGO"),
    ]
    for col, titulo in columnas_base:
        ws.merge_cells(start_row=6, start_column=col, end_row=7, end_column=col)
        celda = ws.cell(row=6, column=col, value=titulo)
        celda.font = fuente_negrita
        celda.alignment = alineacion_centro
        celda.fill = fondo_gris
        celda.border = borde_fino
        ws.cell(row=7, column=col).border = borde_fino

    # Sección SALDOS
    rango_saldo_ini = f"{get_column_letter(COL_SALDO_INI)}6:{get_column_letter(COL_SALDO_INI)}7"
    ws.merge_cells(rango_saldo_ini)
    ws.cell(row=6, column=COL_SALDO_INI, value="CANTIDAD\nSALDOS INICIO\nDEL PERIODO")
    estilizar_rango(rango_saldo_ini, fuente_negrita, alineacion_centro, fondo_gris, borde_fino)

    # INGRESOS
    rango_ing = f"{get_column_letter(COL_ING_START)}6:{get_column_letter(COL_ING_END)}6"
    ws.merge_cells(rango_ing)
    ws.cell(row=6, column=COL_ING_START, value="INGRESOS")
    estilizar_rango(rango_ing, fuente_negrita, alineacion_centro, fondo_gris, borde_fino)
    for col in range(COL_ING_START, COL_ING_END + 1, 2):
        ws.cell(row=7, column=col, value="FECHA").font = fuente_negrita
        ws.cell(row=7, column=col).alignment = alineacion_centro
        ws.cell(row=7, column=col).fill = fondo_gris
        ws.cell(row=7, column=col).border = borde_fino
        ws.cell(row=7, column=col + 1, value="CANTIDAD").font = fuente_negrita
        ws.cell(row=7, column=col + 1).alignment = alineacion_centro
        ws.cell(row=7, column=col + 1).fill = fondo_gris
        ws.cell(row=7, column=col + 1).border = borde_fino

    # TOTAL INGRESOS
    rango_tot_ing = f"{get_column_letter(COL_TOTAL_ING)}6:{get_column_letter(COL_TOTAL_ING)}7"
    ws.merge_cells(rango_tot_ing)
    ws.cell(row=6, column=COL_TOTAL_ING, value="TOTAL\nINGRESOS")
    estilizar_rango(rango_tot_ing, fuente_negrita, alineacion_centro, fondo_gris, borde_fino)

    # EGRESOS
    rango_egr = f"{get_column_letter(COL_EGR_START)}6:{get_column_letter(COL_EGR_END)}6"
    ws.merge_cells(rango_egr)
    ws.cell(row=6, column=COL_EGR_START, value="EGRESOS")
    estilizar_rango(rango_egr, fuente_negrita, alineacion_centro, fondo_gris, borde_fino)
    for col in range(COL_EGR_START, COL_EGR_END + 1, 2):
        ws.cell(row=7, column=col, value="FECHA").font = fuente_negrita
        ws.cell(row=7, column=col).alignment = alineacion_centro
        ws.cell(row=7, column=col).fill = fondo_gris
        ws.cell(row=7, column=col).border = borde_fino
        ws.cell(row=7, column=col + 1, value="CANTIDAD").font = fuente_negrita
        ws.cell(row=7, column=col + 1).alignment = alineacion_centro
        ws.cell(row=7, column=col + 1).fill = fondo_gris
        ws.cell(row=7, column=col + 1).border = borde_fino

    # TOTAL EGRESOS
    rango_tot_egr = f"{get_column_letter(COL_TOTAL_EGR)}6:{get_column_letter(COL_TOTAL_EGR)}7"
    ws.merge_cells(rango_tot_egr)
    ws.cell(row=6, column=COL_TOTAL_EGR, value="TOTAL\nEGRESOS")
    estilizar_rango(rango_tot_egr, fuente_negrita, alineacion_centro, fondo_gris, borde_fino)

    # SALDO FINAL
    rango_saldo_fin = f"{get_column_letter(COL_SALDO_FIN)}6:{get_column_letter(COL_SALDO_FIN)}7"
    ws.merge_cells(rango_saldo_fin)
    ws.cell(row=6, column=COL_SALDO_FIN, value="SALDOS AL FINAL\nDEL PERIODO")
    estilizar_rango(rango_saldo_fin, fuente_negrita, alineacion_centro, fondo_gris, borde_fino)

    # VERIFICACION
    rango_verif = f"{get_column_letter(COL_VERIFICACION)}6:{get_column_letter(COL_VERIFICACION)}7"
    ws.merge_cells(rango_verif)
    ws.cell(row=6, column=COL_VERIFICACION, value="VERIFICACION EXISTENCIAS\nEN FISICO")
    estilizar_rango(rango_verif, fuente_negrita, alineacion_centro, fondo_gris, borde_fino)

    # ============================================================
    # ANCHOS DE COLUMNA
    # ============================================================
    anchos = {
        'A': 5, 'B': 35, 'C': 16, 'D': 10, 'E': 16, 'F': 20, 'G': 20,
        'H': 14, 'I': 14, 'J': 12, 'K': 14,
    }
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho
    ws.column_dimensions[get_column_letter(COL_SALDO_INI)].width = 12
    ws.column_dimensions[get_column_letter(COL_TOTAL_ING)].width = 10
    ws.column_dimensions[get_column_letter(COL_TOTAL_EGR)].width = 10
    ws.column_dimensions[get_column_letter(COL_SALDO_FIN)].width = 12
    ws.column_dimensions[get_column_letter(COL_VERIFICACION)].width = 14

    # ============================================================
    # DATOS
    # ============================================================
    fila_actual = 8
    for idx, data in enumerate(datos_kardex, start=1):
        ws.row_dimensions[fila_actual].height = 20
        m = data['stock'].medicamento
        stock = data['stock']

        # Clasificación de riesgo según INVIMA: I, IIa, IIb, III
        clasif = getattr(m, 'clasificacion_riesgo', '')

        semaforo = calcular_semaforo(stock.fecha_vencimiento)

        datos_fila = {
            COL_ITEM: idx,
            COL_DESCRIPCION: f"{m.principio_activo} {m.concentracion or ''}".strip(),
            COL_MARCA: getattr(m, 'laboratorio', ''),
            COL_SERIE: '',
            COL_PRESENTACION: getattr(m, 'presentacion', ''),
            COL_REG_INVIMA: m.registro_invima or '',
            COL_VIDA_UTIL: getattr(m, 'vida_util', ''),
            COL_LOTE: stock.lote,
            COL_FECHA_VENCE: stock.fecha_vencimiento.strftime('%Y-%m-%d') if stock.fecha_vencimiento else '',
            COL_SEMAFORIZACION: semaforo,
            COL_CLASIF_RIESGO: clasif,
        }

        for col, valor in datos_fila.items():
            celda = ws.cell(row=fila_actual, column=col, value=valor)
            celda.font = fuente_normal
            celda.alignment = alineacion_centro if col not in [COL_DESCRIPCION, COL_PRESENTACION] else alineacion_izq
            celda.border = borde_fino

        # Colorear celda de semáforo
        colores_semaforo = {
            'VENCIDO': 'FF0000', 'ROJO': 'FF4444',
            'NARANJA': 'FF8800', 'AMARILLO': 'FFDD00',
            'VERDE': '44BB44'
        }
        color_hex = colores_semaforo.get(semaforo, '')
        if color_hex:
            ws.cell(row=fila_actual, column=COL_SEMAFORIZACION).fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type='solid'
            )
            ws.cell(row=fila_actual, column=COL_SEMAFORIZACION).font = Font(
                bold=True, size=9, name='Arial',
                color='FFFFFF' if semaforo in ('VENCIDO', 'ROJO', 'NARANJA', 'VERDE') else '000000'
            )

        # Saldo Inicial
        ws.cell(row=fila_actual, column=COL_SALDO_INI, value=data['saldo_inicial']).font = fuente_normal
        ws.cell(row=fila_actual, column=COL_SALDO_INI).alignment = alineacion_centro
        ws.cell(row=fila_actual, column=COL_SALDO_INI).border = borde_fino

        # Ingresos dinámicos
        c_ing = COL_ING_START
        for mov in data['ingresos']:
            ws.cell(row=fila_actual, column=c_ing,
                    value=mov.documento.fecha.strftime("%d/%m")).alignment = alineacion_centro
            ws.cell(row=fila_actual, column=c_ing).font = fuente_normal
            ws.cell(row=fila_actual, column=c_ing).border = borde_fino
            ws.cell(row=fila_actual, column=c_ing + 1, value=mov.cantidad).alignment = alineacion_centro
            ws.cell(row=fila_actual, column=c_ing + 1).font = fuente_normal
            ws.cell(row=fila_actual, column=c_ing + 1).border = borde_fino
            c_ing += 2
        for c in range(COL_ING_START, COL_ING_END + 1):
            if ws.cell(row=fila_actual, column=c).value is None:
                ws.cell(row=fila_actual, column=c).border = borde_fino

        ws.cell(row=fila_actual, column=COL_TOTAL_ING, value=data['total_ingresos']).alignment = alineacion_centro
        ws.cell(row=fila_actual, column=COL_TOTAL_ING).font = fuente_negrita
        ws.cell(row=fila_actual, column=COL_TOTAL_ING).border = borde_fino

        c_eg = COL_EGR_START
        for mov in data['egresos']:
            ws.cell(row=fila_actual, column=c_eg,
                    value=mov.documento.fecha.strftime("%d/%m")).alignment = alineacion_centro
            ws.cell(row=fila_actual, column=c_eg).font = fuente_normal
            ws.cell(row=fila_actual, column=c_eg).border = borde_fino
            ws.cell(row=fila_actual, column=c_eg + 1, value=mov.cantidad).alignment = alineacion_centro
            ws.cell(row=fila_actual, column=c_eg + 1).font = fuente_normal
            ws.cell(row=fila_actual, column=c_eg + 1).border = borde_fino
            c_eg += 2
        for c in range(COL_EGR_START, COL_EGR_END + 1):
            if ws.cell(row=fila_actual, column=c).value is None:
                ws.cell(row=fila_actual, column=c).border = borde_fino

        ws.cell(row=fila_actual, column=COL_TOTAL_EGR, value=data['total_egresos']).alignment = alineacion_centro
        ws.cell(row=fila_actual, column=COL_TOTAL_EGR).font = fuente_negrita
        ws.cell(row=fila_actual, column=COL_TOTAL_EGR).border = borde_fino

        ws.cell(row=fila_actual, column=COL_SALDO_FIN, value=data['saldo_final']).alignment = alineacion_centro
        ws.cell(row=fila_actual, column=COL_SALDO_FIN).font = fuente_negrita
        ws.cell(row=fila_actual, column=COL_SALDO_FIN).border = borde_fino

        ws.cell(row=fila_actual, column=COL_VERIFICACION).border = borde_fino

        for c in range(1, COL_SALDO_INI):
            if ws.cell(row=fila_actual, column=c).border is None or not ws.cell(row=fila_actual, column=c).border.left.style:
                ws.cell(row=fila_actual, column=c).border = borde_fino

        fila_actual += 1

    # ============================================================
    # FILA FINAL: Responsable
    # ============================================================
    fila_actual += 1
    ws.cell(row=fila_actual, column=1, value="NOMBRE DEL RESPONSABLE:").font = fuente_negrita
    ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=5)
    ws.cell(row=fila_actual, column=2).font = fuente_normal

    ws.cell(row=fila_actual, column=7, value="FIRMA DEL RESPONSABLE:").font = fuente_negrita
    ws.cell(row=fila_actual, column=10, value="FECHA:").font = fuente_negrita
    ws.cell(row=fila_actual, column=11, value=hoy.strftime("%d/%m/%Y")).font = fuente_normal

    ws.freeze_panes = 'A8'


def generar_excel_dispositivos(mes, anio, ubicacion_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "CONTROL DE DM E INSUMOS"
    _write_dispositivos_kardex(ws, mes, anio, ubicacion_id)
    return wb


def generar_excel_kardex_consolidado(mes, anio, tipo='MEDICAMENTO'):
    """
    Genera un workbook con una pestaña (sheet) por cada sede.
    Cada pestaña contiene el kardex completo de esa sede.
    """
    wb = Workbook()
    default_ws = wb.active
    sedes = Ubicacion.objects.all().order_by('nombre')

    for idx, sede in enumerate(sedes):
        if idx == 0:
            ws = default_ws
        else:
            ws = wb.create_sheet()
        # Excel sheet names max 31 chars
        ws.title = sede.nombre[:31]

        if tipo == 'DISPOSITIVO':
            _write_dispositivos_kardex(ws, mes, anio, sede.id)
        else:
            _write_medicamentos_kardex(ws, mes, anio, sede.id)

    return wb


def procesar_carga_masiva_productos(usuario, archivo_csv):
    """
    Carga masiva desde CSV. La plantilla incluye todos los campos.
    Según el tipo (MEDICAMENTO|DISPOSITIVO) se toman solo los relevantes.

    Columnas completas:
      tipo, principio_activo, forma_farmaceutica, concentracion,
      presentacion, laboratorio, codigo, registro_invima,
      vida_util, clasificacion_riesgo,
      lote, fecha_vencimiento, cantidad, stock_minimo
    """
    def _normalizar_fecha(texto):
        """Convierte fechas en múltiples formatos a YYYY-MM-DD"""
        if not texto:
            return ''
        texto = texto.strip()
        # Si ya está en formato ISO
        if re.match(r'^\d{4}-\d{2}-\d{2}$', texto):
            return texto
        # DD/MM/YYYY o D/M/YYYY
        match = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', texto)
        if match:
            d, m, y = match.groups()
            return f'{y}-{int(m):02d}-{int(d):02d}'
        # DD-MM-YYYY
        match = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', texto)
        if match:
            d, m, y = match.groups()
            return f'{y}-{int(m):02d}-{int(d):02d}'
        # MM/DD/YYYY (formato US, poco comun)
        match = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', texto)
        if match:
            # Asumir DD/MM/YYYY
            d, m, y = match.groups()
            return f'{y}-{int(m):02d}-{int(d):02d}'
        return texto

    if not usuario.groups.filter(name='ADMIN').exists():
        raise PermissionError("No tienes permisos para realizar cargas masivas.")

    decoded_file = archivo_csv.read().decode('utf-8')
    lines = decoded_file.split('\n')

    # Buscar la línea de encabezado real (la que empieza con 'tipo')
    header_idx = None
    for i, line in enumerate(lines):
        if line.strip().lower().startswith('tipo'):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError('No se encontró la fila de encabezados (debe comenzar con "tipo").')

    # Reconstruir el CSV desde el encabezado real
    csv_body = '\n'.join(lines[header_idx:])
    reader = csv.DictReader(io.StringIO(csv_body))

    # Normalizar nombres de columnas: 'cum' → 'codigo'
    if reader.fieldnames and 'cum' in reader.fieldnames and 'codigo' not in reader.fieldnames:
        reader.fieldnames = ['codigo' if f == 'cum' else f for f in reader.fieldnames]

    # Determinar bodega principal
    bodega_principal = Ubicacion.objects.filter(es_bodega_principal=True).first()
    if not bodega_principal:
        raise ValueError('No hay una bodega principal configurada (es_bodega_principal=True).')

    # Cache de sedes para resolver nombres
    cache_sedes = {}
    def _resolver_sede(nombre):
        if not nombre:
            return None
        key = nombre.strip().lower()
        if key not in cache_sedes:
            cache_sedes[key] = Ubicacion.objects.filter(nombre__iexact=nombre.strip()).first()
        return cache_sedes[key]

    contador = 0
    entradas = []  # (medicamento_id, lote, cantidad)
    traslados_pendientes = []  # (medicamento_id, lote, cantidad, sede_destino)

    with transaction.atomic():
        for row in reader:
            tipo = row.get('tipo', 'MEDICAMENTO').strip().upper()
            if tipo not in ('MEDICAMENTO', 'DISPOSITIVO'):
                tipo = 'MEDICAMENTO'

            principio = row.get('principio_activo', '').strip().upper()

            cups_codigo = row.get('cups_codigo', '').strip()
            if not cups_codigo:
                # Fallback: usar codigo (CUM) como CUPS
                cups_codigo = row.get('codigo', '').strip()
            if not cups_codigo:
                cups_codigo = None

            if tipo == 'DISPOSITIVO':
                forma = 'NO APLICA'
                medicamento, _ = Medicamento.objects.get_or_create(
                    principio_activo=principio,
                    forma_farmaceutica=forma
                )
            else:
                forma = row.get('forma_farmaceutica', row.get('forma', '')).strip().upper()
                medicamento, _ = Medicamento.objects.get_or_create(
                    principio_activo=principio,
                    forma_farmaceutica=forma
                )

            # --- Campos comunes ---
            medicamento.tipo = tipo
            codigo_val = row.get('codigo', '').strip() or None
            # Si el codigo ya está en uso por otro medicamento, no forzar
            if codigo_val and Medicamento.objects.filter(codigo=codigo_val).exclude(id=medicamento.id).exists():
                codigo_val = None
            medicamento.codigo = codigo_val
            medicamento.presentacion = row.get('presentacion', '')
            medicamento.laboratorio = row.get('laboratorio', '')
            medicamento.registro_invima = row.get('registro_invima', '')

            medicamento.cups_codigo = cups_codigo

            if tipo == 'MEDICAMENTO':
                medicamento.concentracion = row.get('concentracion', '')
            else:
                medicamento.concentracion = ''
                medicamento.vida_util = row.get('vida_util', '')
                medicamento.clasificacion_riesgo = row.get('clasificacion_riesgo', '').strip()

            medicamento.save()

            # --- Stock en BODEGA PRINCIPAL ---
            lote = row.get('lote', '').strip().upper()
            if not lote:
                continue  # saltar filas sin lote

            fecha_vto = _normalizar_fecha(row.get('fecha_vencimiento', row.get('vencimiento', '')))
            cantidad = int(row.get('cantidad', 0) or 0)
            stock_min = int(row.get('stock_minimo') or 10)

            stock, _ = InventarioStock.objects.get_or_create(
                ubicacion=bodega_principal,
                medicamento=medicamento,
                lote=lote,
                defaults={
                    'fecha_vencimiento': fecha_vto,
                    'cantidad_actual': 0,
                    'stock_minimo': stock_min
                }
            )

            stock.cantidad_actual += cantidad
            stock.save()

            entradas.append((medicamento.id, lote, cantidad))

            # Leer sede de la fila (columna 'sede' en el CSV)
            sede_nombre = row.get('sede', '').strip()
            sede_destino = _resolver_sede(sede_nombre) if sede_nombre else None
            if sede_destino and sede_destino.id != bodega_principal.id:
                traslados_pendientes.append((medicamento.id, lote, cantidad, sede_destino))

            contador += 1

        # Registrar un solo documento ENTRADA con todos los detalles
        if entradas:
            doc_entrada = Documento.objects.create(
                tipo_mov='ENTRADA',
                destino=bodega_principal,
                usuario=usuario,
            )
            for med_id, lote, cantidad in entradas:
                DocumentoDetalle.objects.create(
                    documento=doc_entrada,
                    medicamento_id=med_id,
                    lote=lote,
                    cantidad=cantidad,
                )

        # Ejecutar traslados pendientes (bodega → sede indicada en la fila)
        if traslados_pendientes:
            # Agrupar por sede destino para crear un documento TRASLADO por sede
            traslados_por_sede = {}
            for med_id, lote, cantidad, sede_dest in traslados_pendientes:
                sede_key = sede_dest.id
                if sede_key not in traslados_por_sede:
                    traslados_por_sede[sede_key] = {
                        'sede': sede_dest,
                        'items': []
                    }
                traslados_por_sede[sede_key]['items'].append((med_id, lote, cantidad))

            for sede_key, grupo in traslados_por_sede.items():
                sede_dest = grupo['sede']
                items = grupo['items']

                doc_traslado = Documento.objects.create(
                    tipo_mov='TRASLADO',
                    origen=bodega_principal,
                    destino=sede_dest,
                    usuario=usuario,
                )
                for med_id, lote, cantidad in items:
                    stock_origen = InventarioStock.objects.select_for_update().get(
                        ubicacion=bodega_principal, medicamento_id=med_id, lote=lote
                    )
                    if stock_origen.cantidad_actual < cantidad:
                        continue
                    stock_destino, _ = InventarioStock.objects.get_or_create(
                        ubicacion=sede_dest,
                        medicamento_id=med_id,
                        lote=lote,
                        defaults={
                            'cantidad_actual': 0,
                            'fecha_vencimiento': stock_origen.fecha_vencimiento,
                            'stock_minimo': stock_origen.stock_minimo or 10,
                        }
                    )
                    stock_origen.cantidad_actual -= cantidad
                    stock_origen.save()
                    stock_destino.cantidad_actual += cantidad
                    stock_destino.save()
                    DocumentoDetalle.objects.create(
                        documento=doc_traslado,
                        medicamento_id=med_id,
                        lote=lote,
                        cantidad=cantidad,
                    )

    return contador


def registrar_solicitud_reabastecimiento(usuario, nombre_med, cantidad):
    """Crea un documento de tipo SOLICITUD para que el admin lo vea"""
    return Documento.objects.create(
        tipo_mov='SOLICITUD',
        usuario=usuario,
        origen=usuario.perfil.ubicacion_asignada,
        id_paciente=f"REQ-{nombre_med[:3].upper()}",  # Tag de seguimiento
        # En los detalles guardamos qué se pidió
    )


def generar_plantilla_xlsx():
    """
    Genera un workbook XLSX con la plantilla de carga masiva.
    Incluye dos filas de ejemplo: MEDICAMENTO y DISPOSITIVO.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Carga Masiva"

    fuente_titulo = Font(bold=True, size=11, name='Arial', color='FFFFFF')
    fuente_normal = Font(size=10, name='Arial')
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alineacion_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borde_fino = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fondo_azul = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    fondo_ejemplo_med = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    fondo_ejemplo_disp = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

    encabezados = [
        "tipo", "principio_activo", "forma_farmaceutica", "concentracion",
        "presentacion", "laboratorio", "codigo", "cups_codigo",
        "registro_invima", "vida_util", "clasificacion_riesgo",
        "sede",
        "lote", "fecha_vencimiento", "cantidad", "stock_minimo"
    ]

    # Ejemplo medicamento
    ejemplo_med = [
        "MEDICAMENTO", "ACETAMINOFEN", "TABLETA", "500MG",
        "CAJA X 10", "GENFAR", "770123456", "70005",
        "2020M-0012345", "", "",
        "Puerto Tejada",
        "LOTE001", "2026-12-31", "100", "20"
    ]

    # Ejemplo dispositivo
    ejemplo_disp = [
        "DISPOSITIVO", "GUANTES QUIRURGICOS", "NO APLICA", "",
        "CAJA X 100", "ECOPIEL", "", "70905",
        "INVIMA-2025E-001234", "3 AÑOS", "I",
        "",
        "LOTE002", "2027-06-30", "500", "50"
    ]

    # Hoja informativa (fila 1)
    ws.merge_cells('A1:O1')
    ws['A1'] = "PLANTILLA DE CARGA MASIVA - KARDEX FARMACIA"
    ws['A1'].font = Font(bold=True, size=14, name='Arial', color='1E3A8A')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:O2')
    ws['A2'] = "Complete los datos. Todas las columnas con HEADER AZUL son obligatorias. Borre las filas de ejemplo antes de importar."
    ws['A2'].font = Font(size=9, name='Arial', italic=True, color='64748B')
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 25

    # Fila de headers (fila 4)
    row_header = 4
    ws.row_dimensions[row_header].height = 28
    for col_idx, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=row_header, column=col_idx, value=titulo)
        celda.font = fuente_titulo
        celda.alignment = alineacion_centro
        celda.fill = fondo_azul
        celda.border = borde_fino

    # Ejemplo MEDICAMENTO (fila 5)
    row_med = 5
    ws.row_dimensions[row_med].height = 22
    cols_centradas = {1, 4, 12, 13, 14, 15}
    for col_idx, valor in enumerate(ejemplo_med, start=1):
        celda = ws.cell(row=row_med, column=col_idx, value=valor)
        celda.font = fuente_normal
        celda.alignment = alineacion_centro if col_idx in cols_centradas else alineacion_izq
        celda.fill = fondo_ejemplo_med
        celda.border = borde_fino

    # Ejemplo DISPOSITIVO (fila 6)
    row_disp = 6
    ws.row_dimensions[row_disp].height = 22
    for col_idx, valor in enumerate(ejemplo_disp, start=1):
        celda = ws.cell(row=row_disp, column=col_idx, value=valor)
        celda.font = fuente_normal
        celda.alignment = alineacion_centro if col_idx in cols_centradas else alineacion_izq
        celda.fill = fondo_ejemplo_disp
        celda.border = borde_fino

    # Anchos de columna
    anchos = {'A': 14, 'B': 30, 'C': 20, 'D': 14, 'E': 18, 'F': 16,
              'G': 14, 'H': 16, 'I': 22, 'J': 10, 'K': 20,
              'L': 20, 'M': 12, 'N': 18, 'O': 10, 'P': 12}
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    # Congelar paneles
    ws.freeze_panes = 'A5'

    return wb


# ==============================================================================
# SERVICIOS DE CONCILIACIÓN RIPS
# ==============================================================================

import io as _io
import csv as _csv
import datetime as _dt
import re as _re
from django.utils import timezone as _tz
from django.db import transaction as _tx
from django.db.models import Sum as _Sum
from .models import (
    CargaRIPS as _CargaRIPS, RegistroRIPS as _RegistroRIPS,
    Conciliacion as _Conciliacion, DetalleConciliacion as _DetalleConciliacion,
    MapeoRIPSMedicamento as _Mapeo, Medicamento as _Medicamento,
    Documento as _Documento, DocumentoDetalle as _DocDetalle
)


_GRUPOS_MED = frozenset({'MEDICAMENTO', 'SERVICIO FARMACEUTICO', 'INSUMOS'})


def _parse_decimal(v):
    try:
        return float(str(v).replace(',', '.'))
    except (ValueError, AttributeError):
        return 0


def procesar_importacion_rips(archivo_csv, periodo_inicio=None, periodo_fin=None, sede_filter=''):
    """
    Procesa un archivo CSV reporte201 y guarda en RegistroRIPS.
    Retorna: (CargaRIPS, mensaje, success)
    """
    if hasattr(archivo_csv, 'read'):
        content = archivo_csv.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig')
        io_string = _io.StringIO(content)
    else:
        io_string = _io.StringIO(archivo_csv)

    reader = _csv.DictReader(io_string, delimiter='\t')
    if not reader.fieldnames:
        return None, 'El CSV no tiene cabeceras o el delimitador no es tab', False

    if not periodo_inicio:
        periodo_inicio = _tz.now().date().replace(day=1)
    if not periodo_fin:
        periodo_fin = _tz.now().date()

    nombre_archivo = getattr(archivo_csv, 'name', 'upload_report_rips.csv')
    if hasattr(archivo_csv, 'name'):
        nombre_archivo = archivo_csv.name

    with _tx.atomic():
        carga = _CargaRIPS.objects.create(
            archivo=nombre_archivo.replace('\\', '/').split('/')[-1],
            periodo_inicio=periodo_inicio,
            periodo_fin=periodo_fin,
        )

        total = medicamentos = 0
        registros_batch = []
        BATCH = 1000

        for row in reader:
            total += 1
            gruposervicio = row.get('gruposervicio', '').strip().upper()
            sede = row.get('sedehabilitacion', '').strip()

            if sede_filter and sede_filter.upper() not in sede.upper():
                continue
            if gruposervicio not in _GRUPOS_MED:
                continue

            fecha_str = row.get('fechaprocedimiento', '').strip()
            fecha = None
            if fecha_str:
                try:
                    na = _dt.datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
                    fecha = _tz.make_aware(na, _dt.timezone(_dt.timedelta(hours=-5)))
                except ValueError:
                    try:
                        fecha = _dt.datetime.strptime(fecha_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass

            registros_batch.append(_RegistroRIPS(
                carga=carga,
                gruposervicio=gruposervicio,
                cupscodigo=row.get('cupscodigo', '').strip(),
                nombreprocedimiento=row.get('nombreprocedimiento', '').strip(),
                cantidad=int(row.get('cantidad', 0) or 0),
                valorunitario=_parse_decimal(row.get('valorunidad', '0')),
                valortotal=_parse_decimal(row.get('valortotal', '0')),
                identificacion_paciente=row.get('identificacion', '').strip(),
                nombre_paciente=row.get('nombrecompleto', '').strip(),
                identificacion_profesional=row.get('identificacionprofesional', '').strip(),
                nombre_profesional=row.get('nombreprofesional', '').strip(),
                especialidad=row.get('especialidad', '').strip(),
                fechaprocedimiento=fecha,
                numerofactura=row.get('numerofactura', '').strip(),
                sede=sede,
                admision=row.get('admision', '').strip(),
                modalidad=row.get('modalidad', '').strip(),
                diagnostico=row.get('diagnostico', '').strip(),
                diagnosticonombre=row.get('diagnosticonombre', '').strip(),
            ))
            medicamentos += 1

            if len(registros_batch) >= BATCH:
                _RegistroRIPS.objects.bulk_create(registros_batch, ignore_conflicts=True)
                registros_batch = []

        if registros_batch:
            _RegistroRIPS.objects.bulk_create(registros_batch, ignore_conflicts=True)

        carga.total_registros = total
        carga.registros_medicamentos = medicamentos

        if medicamentos == 0:
            carga.estado = 'ERROR'
            carga.save()
            return carga, 'No se encontraron registros de medicamentos en el archivo', False

        carga.save()

    _mapear_medicamentos_rips(carga)
    return carga, f'Cargados {medicamentos} registros de medicamentos/insumos (de {total} totales)', True


def _mapear_medicamentos_rips(carga):
    """
    Mapea registros RIPS a medicamentos del Kardex en este orden:
    1. Por cups_codigo directo del Medicamento (exacto contra cupscodigo del RIPS)
    2. Por MapeoRIPSMedicamento explícito (cups_codigo + nombre_rips + gruposervicio)
    3. Por texto: principio_activo dentro del nombre del procedimiento RIPS
    """
    # 1. Matching directo por CUPS code en el medicamento
    for med in _Medicamento.objects.exclude(cups_codigo__isnull=True).exclude(cups_codigo__exact=''):
        _RegistroRIPS.objects.filter(
            carga=carga, cupscodigo=med.cups_codigo,
            medicamento_mapeado__isnull=True
        ).update(medicamento_mapeado=med)

    # 2. Mapeos explícitos
    mapeos = _Mapeo.objects.filter(activo=True)
    for m in mapeos:
        f = {}
        if m.cups_codigo:
            f['cupscodigo'] = m.cups_codigo
        if m.nombre_rips:
            f['nombreprocedimiento__iexact'] = m.nombre_rips
        if m.gruposervicio:
            f['gruposervicio__iexact'] = m.gruposervicio
        if f:
            f['medicamento_mapeado__isnull'] = True
            _RegistroRIPS.objects.filter(carga=carga, **f).update(medicamento_mapeado=m.medicamento)

    # 3. Matching por principio activo en texto
    for r in _RegistroRIPS.objects.filter(carga=carga, medicamento_mapeado__isnull=True):
        n = r.nombreprocedimiento.upper()
        for m in _Medicamento.objects.all():
            if m.principio_activo.upper() in n:
                r.medicamento_mapeado = m
                r.save(update_fields=['medicamento_mapeado'])
                break


def ejecutar_conciliacion(carga):
    """Ejecuta conciliación Kardex vs RIPS y retorna la Conciliacion creada"""
    pi, pf = carga.periodo_inicio, carga.periodo_fin

    salidas = _Documento.objects.filter(
        tipo_mov='SALIDA', fecha__date__gte=pi, fecha__date__lte=pf
    ).select_related('origen', 'usuario').prefetch_related('detalles__medicamento')

    salidas_dict, tot_cant_k = {}, 0
    for doc in salidas:
        for det in doc.detalles.all():
            tot_cant_k += det.cantidad
            k = (doc.id_paciente or '', doc.fecha.date(), det.medicamento.principio_activo.upper())
            salidas_dict.setdefault(k, []).append({'doc': doc, 'det': det, 'cant': det.cantidad})

    rips = _RegistroRIPS.objects.filter(carga=carga)
    tot_rips = rips.count()
    tot_cant_r = rips.aggregate(t=_Sum('cantidad'))['t'] or 0

    rips_dict = {}
    for r in rips:
        f = r.fechaprocedimiento.date() if r.fechaprocedimiento else pi
        rips_dict.setdefault((r.identificacion_paciente, f, r.nombreprocedimiento.upper().strip()), []).append(r)

    with _tx.atomic():
        conc = _Conciliacion.objects.create(
            carga_rips=carga, periodo_inicio=pi, periodo_fin=pf,
            total_salidas_kardex=salidas.count(), total_medicamentos_kardex=tot_cant_k,
            total_registros_rips=tot_rips, total_cantidad_rips=tot_cant_r,
        )

        ok, nf, nd, batch = 0, 0, 0, []

        for (paci, fech, med_nom), items in salidas_dict.items():
            tk = sum(i['cant'] for i in items)
            hallado = False
            for kr, regs in list(rips_dict.items()):
                p, fr, nr = kr
                if paci != p: continue
                if abs((fech - fr).days) > 1: continue
                if med_nom not in nr and nr not in med_nom: continue
                hallado = True
                tr = sum(r.cantidad for r in regs)
                for i in items:
                    ed = 'COINCIDE' if i['cant'] == tr else 'CANTIDAD_DIF'
                    batch.append(_DetalleConciliacion(
                        conciliacion=conc, estado=ed, documento_salida=i['doc'],
                        medicamento_nombre=i['det'].medicamento.principio_activo,
                        paciente_identificacion=paci, paciente_nombre=regs[0].nombre_paciente,
                        cantidad_kardex=i['cant'], cantidad_rips=tr,
                        fecha=i['doc'].fecha,
                        sede=i['doc'].origen.nombre if i['doc'].origen else '',
                        profesional=i['doc'].usuario.get_full_name() or i['doc'].usuario.username,
                        observacion='OK' if i['cant'] == tr else f"K: {i['cant']} vs R: {tr}",
                    ))
                ok += 1 if i['cant'] == tr else 0
                nf += 0 if i['cant'] == tr else 1
                del rips_dict[kr]; break
            if not hallado:
                for i in items:
                    batch.append(_DetalleConciliacion(
                        conciliacion=conc, estado='NO_FACTURADO',
                        documento_salida=i['doc'],
                        medicamento_nombre=i['det'].medicamento.principio_activo,
                        paciente_identificacion=paci, cantidad_kardex=i['cant'],
                        cantidad_rips=0, fecha=i['doc'].fecha,
                        sede=i['doc'].origen.nombre if i['doc'].origen else '',
                        profesional=i['doc'].usuario.get_full_name() or i['doc'].usuario.username,
                        observacion='Dispensado en Kardex pero NO en RIPS',
                    ))
                nf += len(items)

        for kr, regs in rips_dict.items():
            for r in regs:
                batch.append(_DetalleConciliacion(
                    conciliacion=conc, estado='NO_DESPACHADO', registro_rips=r,
                    medicamento_nombre=r.nombreprocedimiento,
                    paciente_identificacion=r.identificacion_paciente,
                    paciente_nombre=r.nombre_paciente,
                    cantidad_kardex=0, cantidad_rips=r.cantidad,
                    fecha=r.fechaprocedimiento or _tz.now(), sede=r.sede,
                    profesional=r.nombre_profesional,
                    observacion='Facturado en RIPS pero NO en Kardex',
                ))
            nd += len(regs)

        if batch:
            _DetalleConciliacion.objects.bulk_create(batch, batch_size=1000)

        conc.coincidencias = ok; conc.no_facturados = nf
        conc.no_despachados = nd; conc.save()

    carga.estado = 'CONCILIADA'; carga.save()
    return conc