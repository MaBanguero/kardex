#!/usr/bin/env python3
"""
Management command para importar auxiliares de enfermería desde un archivo Excel.

Crea:
  - User (username = documento, first_name = nombre, sin last_name)
  - PerfilUsuario (numero_identificacion = documento, must_change_password = True)
  - Asigna al grupo ENFERMERA
  - Clave temporal: el número de documento completo

Uso:
  python3 manage.py importar_enfermeras --archivo=ruta/al/archivo.xlsx
  python3 manage.py importar_enfermeras --archivo=ruta --sede=1
"""

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User, Group
from django.db import transaction
from kardex.models import PerfilUsuario, Ubicacion


class Command(BaseCommand):
    help = 'Importa auxiliares de enfermería desde un archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            type=str,
            required=True,
            help='Ruta al archivo Excel con los datos de enfermeras'
        )
        parser.add_argument(
            '--sede',
            type=int,
            default=None,
            help='ID de la sede/ubicación a asignar (opcional)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Solo mostrar lo que se haría sin modificar la base de datos'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        sede_id = options['sede']
        dry_run = options['dry_run']

        # Validar archivo
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception as e:
            raise CommandError(f'No se pudo leer el archivo: {e}')

        # Validar sede si se especificó
        sede = None
        if sede_id:
            try:
                sede = Ubicacion.objects.get(id=sede_id)
            except Ubicacion.DoesNotExist:
                raise CommandError(f'No existe una sede con ID {sede_id}')

        # Asegurar que existe el grupo ENFERMERA
        grupo_enfermera, _ = Group.objects.get_or_create(name='ENFERMERA')

        # Leer datos del Excel
        registros = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 3:
                continue
            name = str(row[1]).strip() if row[1] else ''
            document = str(row[2]).strip() if row[2] else ''
            document = ''.join(c for c in document if c.isdigit())

            if not name or not document:
                continue
            if name.lower() in ('', 'nombre', 'nombres', 'apellidos', 'none'):
                continue

            # Clave temporal: el documento completo
            temp_pass = document
            username = document  # username = documento

            registros.append({
                'username': username,
                'password': temp_pass,
                'first_name': name,
                'document': document,
                'temp_pass': temp_pass,
            })

        if not registros:
            self.stdout.write(self.style.WARNING('⚠️  No se encontraron datos válidos en el archivo.'))
            return

        self.stdout.write(f'📂 Archivo: {archivo}')
        self.stdout.write(f'📋 Enfermeras encontradas: {len(registros)}')
        if sede:
            self.stdout.write(f'🏥 Sede asignada: {sede.nombre} (ID: {sede.id})')
        else:
            self.stdout.write(self.style.WARNING('⚠️  Sin sede asignada. Las enfermeras deberán tener sede para operar.'))
        self.stdout.write('')

        if dry_run:
            self.stdout.write(self.style.WARNING('🔍 MODO DRY-RUN — No se modifica la base de datos'))
            self.stdout.write('')

        # Mostrar tabla
        self.stdout.write(f'{"#":<4} {"NOMBRE":<35} {"DOCUMENTO":<15} {"CLAVE TEMP":<15}')
        self.stdout.write('-' * 70)
        for i, r in enumerate(registros, 1):
            self.stdout.write(f'{i:<4} {r["first_name"][:34]:<35} {r["document"]:<15} {r["temp_pass"]:<15}')

        self.stdout.write('')

        if dry_run:
            return

        # Importar
        creados = 0
        actualizados = 0
        errores = []

        with transaction.atomic():
            for r in registros:
                try:
                    user, created = User.objects.get_or_create(
                        username=r['username'],
                        defaults={
                            'first_name': r['first_name'],
                            'last_name': '',
                            'email': '',
                        }
                    )

                    if created:
                        user.set_password(r['temp_pass'])
                        user.save()

                        PerfilUsuario.objects.create(
                            usuario=user,
                            ubicacion_asignada=sede,
                            numero_identificacion=r['document'],
                            must_change_password=True,
                        )

                        user.groups.add(grupo_enfermera)
                        creados += 1
                        self.stdout.write(f'  ✅ Creada: {r["first_name"]} ({r["document"]})')
                    else:
                        # Si ya existe, actualizar datos y resetear clave
                        user.first_name = r['first_name']
                        user.set_password(r['temp_pass'])
                        user.save()

                        perfil, _ = PerfilUsuario.objects.get_or_create(
                            usuario=user,
                            defaults={
                                'ubicacion_asignada': sede,
                                'numero_identificacion': r['document'],
                                'must_change_password': True,
                            }
                        )
                        if not perfil.must_change_password:
                            perfil.must_change_password = True
                            perfil.save()

                        user.groups.add(grupo_enfermera)
                        actualizados += 1
                        self.stdout.write(self.style.WARNING(f'  🔄 Actualizada: {r["first_name"]} ({r["document"]})'))

                except Exception as e:
                    errores.append(f'{r["document"]}: {e}')
                    self.stdout.write(self.style.ERROR(f'  ❌ Error con {r["document"]}: {e}'))

        # Resumen final
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=' * 55))
        self.stdout.write(self.style.SUCCESS(f'  ✅ Importación completada'))
        self.stdout.write(self.style.SUCCESS(f'  🆕 Creadas: {creados}'))
        self.stdout.write(self.style.SUCCESS(f'  🔄 Actualizadas: {actualizados}'))
        if errores:
            self.stdout.write(self.style.ERROR(f'  ❌ Errores: {len(errores)}'))
        self.stdout.write(self.style.SUCCESS(f'  🔑 Clave temporal: el número de documento completo'))
        self.stdout.write(self.style.SUCCESS('=' * 55))
