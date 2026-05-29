import uuid
import json
import datetime
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from .models import Documento, ConfiguracionSistema, DocumentoDetalle, Medicamento, User, Ubicacion, SolicitudStock, TurnoEnfermera
from django.db import transaction
from django.db.models import Sum, Q as models_Q, Value as V
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import timedelta

# Importamos nuestros modelos y la lógica ACID
from .models import InventarioStock, PerfilUsuario, Conciliacion
from .services import generar_excel_kardex, generar_excel_kardex_consolidado, calcular_semaforo, registrar_salida_paciente_inteligente, registrar_devolucion_agrupada, procesar_carga_masiva_productos, generar_plantilla_xlsx
from django.contrib.auth.views import LoginView
from django.contrib.auth.models import Group
from django.urls import reverse_lazy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def get_turno_activo(ubicacion):
    """
    Retorna el TurnoEnfermera activo NO expirado para una sede,
    o None si no hay turno activo.
    """
    if not ubicacion:
        return None
    now = timezone.now()
    hora_limite = now - timedelta(hours=12)
    return TurnoEnfermera.objects.filter(
        sede=ubicacion,
        activo=True,
        fecha_inicio__gte=hora_limite,
        fecha_expiracion__gt=now,
    ).select_related('enfermera').first()


class CustomLoginView(LoginView):
    template_name = 'kardex/login.html'
    redirect_authenticated_user = True

    def form_valid(self, form):
        """
        Login exitoso con lógica de turno único para ENFERMERA.
        - Si must_change_password=True: salta la validación de turno (va directo a cambio de clave)
        - Si es ENFERMERA sin must_change: valida turno único (1 activa por sede cada 12h)
        - ADMIN y REGENTE: sin restricción de turno
        """
        user = form.get_user()

        # Verificar que tenga perfil y sede
        try:
            perfil = user.perfil
            if not perfil.ubicacion_asignada:
                form.add_error(None, 'Tu usuario no tiene una sede asignada. Contacta al administrador.')
                return self.form_invalid(form)
        except PerfilUsuario.DoesNotExist:
            form.add_error(None, 'Este usuario no tiene un perfil configurado. Contacta al administrador.')
            return self.form_invalid(form)

        # Si debe cambiar la clave → salta toda validación de turno
        if perfil.must_change_password:
            return super().form_valid(form)

        # Validación de turno único solo para ENFERMERA (con clave ya cambiada)
        grupos = user.groups.values_list('name', flat=True)

        if 'ENFERMERA' in grupos:
            sede = user.perfil.ubicacion_asignada

            now = timezone.now()
            hora_limite = now - timedelta(hours=12)

            # Buscar turnos activos NO expirados de OTRAS enfermeras en la misma sede
            turno_ocupado = TurnoEnfermera.objects.filter(
                sede=sede,
                activo=True,
                fecha_inicio__gte=hora_limite,
                fecha_expiracion__gt=now,
            ).exclude(enfermera=user).first()

            if turno_ocupado:
                nombre_enfermera = turno_ocupado.enfermera.get_full_name() or turno_ocupado.enfermera.username
                form.add_error(
                    None,
                    f'Ya hay un turno activo de {nombre_enfermera} en {sede.nombre}. '
                    f'Debe esperar a que finalice su turno (12h desde su inicio) para iniciar sesión.'
                )
                return self.form_invalid(form)

            # Turno de la MISMA enfermera: renovar expiración
            turno_propio, created = TurnoEnfermera.objects.get_or_create(
                enfermera=user,
                sede=sede,
                defaults={
                    'activo': True,
                    'fecha_expiracion': now + timedelta(hours=12),
                }
            )
            if not created:
                turno_propio.fecha_expiracion = now + timedelta(hours=12)
                turno_propio.activo = True
                turno_propio.save()

        return super().form_valid(form)

    def get_success_url(self):
        """
        Enruta al usuario:
        - Si debe cambiar la contraseña → pantalla de cambio de clave
        - ADMIN/REGENTE → admin_dashboard
        - ENFERMERA → dashboard de usuario
        """
        usuario = self.request.user

        # Forzar cambio de contraseña ANTES de cualquier otra cosa
        if usuario.perfil.must_change_password:
            return reverse_lazy('cambiar_clave')

        grupos_usuario = usuario.groups.values_list('name', flat=True)

        if 'ADMIN' in grupos_usuario or 'REGENTE' in grupos_usuario:
            return reverse_lazy('admin_dashboard')
        elif 'ENFERMERA' in grupos_usuario:
            return reverse_lazy('dashboard')

        return reverse_lazy('dashboard')


# ==========================================
# CAMBIO OBLIGATORIO DE CONTRASEÑA
# ==========================================
from django.contrib import messages
from django.shortcuts import redirect


@login_required
def cambiar_clave_view(request):
    """
    Vista para cambiar la contraseña.
    Si must_change_password=True, fuerza el cambio antes de dejar acceder al sistema.
    """
    perfil = request.user.perfil

    if request.method == 'POST':
        current = request.POST.get('current_password', '')
        new_pass = request.POST.get('new_password', '')
        confirm = request.POST.get('confirm_password', '')

        if not request.user.check_password(current):
            messages.error(request, 'La contraseña actual no es correcta.')
            return render(request, 'kardex/cambiar_clave.html', {'must_change': perfil.must_change_password})

        if not new_pass or not confirm:
            messages.error(request, 'Todos los campos son obligatorios.')
            return render(request, 'kardex/cambiar_clave.html', {'must_change': perfil.must_change_password})

        if new_pass != confirm:
            messages.error(request, 'Las contraseñas nuevas no coinciden.')
            return render(request, 'kardex/cambiar_clave.html', {'must_change': perfil.must_change_password})

        if len(new_pass) < 6:
            messages.error(request, 'La nueva contraseña debe tener al menos 6 caracteres.')
            return render(request, 'kardex/cambiar_clave.html', {'must_change': perfil.must_change_password})

        # Cambiar contraseña
        request.user.set_password(new_pass)
        request.user.save()

        perfil.must_change_password = False
        perfil.last_password_change = timezone.now()
        perfil.save()

        # Re-autenticar para que la sesión no se pierda
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, request.user)

        messages.success(request, '✅ Contraseña cambiada exitosamente. Ya puedes usar el sistema.')
        return redirect('dashboard')

    return render(request, 'kardex/cambiar_clave.html', {'must_change': perfil.must_change_password})


# ==========================================
# 1. VISTA PRINCIPAL (EL DASHBOARD SPA)
# ==========================================
@login_required
def dashboard_kardex(request):
    """
    Renderiza el contenedor principal. Como usamos una arquitectura SPA
    (Single Page Application) con LocalStorage, ya no enviamos todo el 
    inventario aquí. Solo enviamos la ubicación para los títulos.
    """
    try:
        ubicacion_actual = request.user.perfil.ubicacion_asignada

        if not ubicacion_actual:
            return render(request, 'kardex/error.html', {
                'mensaje': 'Tu usuario no tiene una sede/ubicación asignada. Contacta al administrador.'
            })

        # Enfermera de turno activo en esta sede
        turno_activo = get_turno_activo(ubicacion_actual)

        return render(request, 'kardex/dashboard.html', {
            'ubicacion': ubicacion_actual,
            'turno_activo': turno_activo,
        })

    except PerfilUsuario.DoesNotExist:
        return render(request, 'kardex/error.html', {
            'mensaje': 'Este usuario no tiene un perfil clínico configurado. Créalo en el panel de administración.'
        })


# ==========================================
# 2. API PARA EL LOCALSTORAGE (PWA)
# ==========================================
@login_required
def sincronizar_inventario_api(request):
    """Envía el inventario al frontend.
    - ADMIN: ve stock de TODAS las sedes
    - Otros roles: solo su sede asignada
    """
    grupos = request.user.groups.values_list('name', flat=True)
    es_admin = 'ADMIN' in grupos

    if es_admin:
        stock = InventarioStock.objects.all().select_related('medicamento', 'ubicacion')
        meds_pendientes = SolicitudStock.objects.filter(
            estado='PENDIENTE'
        ).values_list('medicamento_id', flat=True)
    else:
        ubicacion = request.user.perfil.ubicacion_asignada
        stock = InventarioStock.objects.filter(ubicacion=ubicacion).select_related('medicamento')
        meds_pendientes = SolicitudStock.objects.filter(
            sede_solicitante=ubicacion,
            estado='PENDIENTE'
        ).values_list('medicamento_id', flat=True)

    data = []
    for item in stock:
        semaforo = calcular_semaforo(item.fecha_vencimiento)
        data.append({
            'id': item.id,
            'medicamento_id': item.medicamento.id,
            'principio_activo': item.medicamento.principio_activo,
            'forma_farmaceutica': item.medicamento.forma_farmaceutica,
            'concentracion': item.medicamento.concentracion,
            'codigo': item.medicamento.codigo,
            'presentacion': item.medicamento.presentacion,
            'laboratorio': item.medicamento.laboratorio,
            'lote': item.lote,
            'fecha_vencimiento': item.fecha_vencimiento.strftime('%Y-%m-%d') if item.fecha_vencimiento else '',
            'cantidad_actual': item.cantidad_actual,
            'stock_minimo': item.stock_minimo,
            'tipo': item.medicamento.tipo,
            'semaforo': semaforo,
            'vida_util': item.medicamento.vida_util or '',
            'clasificacion_riesgo': item.medicamento.clasificacion_riesgo or '',
            'cups_codigo': item.medicamento.cups_codigo or '',
            'ubicacion_id': item.ubicacion_id,
            'ubicacion_nombre': item.ubicacion.nombre if hasattr(item, 'ubicacion') and item.ubicacion else '',
            'busqueda': f"{item.medicamento.principio_activo} {item.lote} {item.medicamento.cups_codigo or ''} {item.medicamento.codigo or ''} {item.ubicacion.nombre if hasattr(item, 'ubicacion') and item.ubicacion else ''}".lower(),
            'en_tramite': item.medicamento.id in meds_pendientes
        })
    response = JsonResponse({'inventario': data})
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    return response


# ==========================================
# 3. PROCESAMIENTO DE MOVIMIENTOS (AJAX)
# ==========================================
@login_required
@require_POST
def registrar_movimiento_view(request):
    try:
        data = json.loads(request.body)
        tipo = data.get('tipo_mov')
        cantidad = int(data.get('cantidad', 0))
        id_paciente = data.get('id_paciente')

        if tipo == 'SALIDA':
            nombre_med = data.get('nombre_medicamento')
            cups_codigo = data.get('cups_codigo')
            presentacion = data.get('presentacion')
            registrar_salida_paciente_inteligente(
                request.user, nombre_med, cantidad, id_paciente,
                cups_codigo=cups_codigo, presentacion=presentacion
            )
            return JsonResponse({'status': 'success', 'requiere_sincronizacion': True})

        elif tipo == 'DEVOLUCION':
            doc_id = data.get('doc_id')
            nombre_med = data.get('nombre_medicamento')

            if doc_id:
                # Devolución específica contra un documento de salida concreto
                from .services import registrar_devolucion
                registrar_devolucion(request.user, doc_id, cantidad)
            else:
                # Devolución agrupada: busca todas las salidas del paciente para este medicamento
                registrar_devolucion_agrupada(request.user, nombre_med, cantidad, id_paciente)

            return JsonResponse({
                'status': 'success',
                'mensaje': f'Se han reingresado {cantidad} unidades al inventario.'
            })

    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


# ==========================================
# 4. EXPORTACIÓN DE REPORTE EXCEL
# ==========================================
@login_required
def exportar_kardex_excel(request):
    """
    Genera el kárdex en formato Excel.
    Parámetros GET:
    - tipo: MEDICAMENTO (default) | DISPOSITIVO
    - sede: ID numérico | 'consolidado' | vacío (usa sede del usuario)
    """
    hoy = datetime.datetime.now()
    tipo = request.GET.get('tipo', 'MEDICAMENTO')
    sede_param = request.GET.get('sede', '')

    grupos = request.user.groups.values_list('name', flat=True)
    es_admin = 'ADMIN' in grupos

    # Modo consolidado: una pestaña por sede (solo admin)
    if sede_param == 'consolidado':
        if not es_admin:
            return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)
        wb = generar_excel_kardex_consolidado(hoy.month, hoy.year, tipo)
        tipo_label = 'Medicamentos' if tipo == 'MEDICAMENTO' else 'Dispositivos'
        filename = f'Kardex_Consolidado_{tipo_label}_{hoy.strftime("%Y%m%d")}.xlsx'

    # Modo sede específica (admin puede elegir cualquier sede)
    elif sede_param and sede_param.isdigit():
        sede_id = int(sede_param)
        if not es_admin and sede_id != request.user.perfil.ubicacion_asignada.id:
            return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)
        wb = generar_excel_kardex(hoy.month, hoy.year, sede_id, tipo)
        tipo_label = 'Medicamentos' if tipo == 'MEDICAMENTO' else 'Dispositivos'
        filename = f'Kardex_{tipo_label}_{hoy.strftime("%Y%m%d")}.xlsx'

    # Modo default: sede del usuario
    else:
        ubicacion_id = request.user.perfil.ubicacion_asignada.id
        wb = generar_excel_kardex(hoy.month, hoy.year, ubicacion_id, tipo)
        tipo_label = 'Medicamentos' if tipo == 'MEDICAMENTO' else 'Dispositivos'
        filename = f'Kardex_{tipo_label}_{hoy.strftime("%Y%m%d")}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response


@login_required
def historial_movimientos_api(request):
    config = ConfiguracionSistema.objects.first()
    limite_horas = config.horas_limite_devolucion if config else 2

    hace_24h = timezone.now() - timedelta(hours=24)
    limite_devolucion = timezone.now() - timedelta(hours=limite_horas)

    # Obtenemos todas las salidas del usuario en las últimas 24h
    movimientos = Documento.objects.filter(
        usuario=request.user,
        tipo_mov='SALIDA',
        fecha__gte=hace_24h
    ).order_by('-fecha')

    historial = []
    for mov in movimientos:
        # Sumamos la cantidad total de este documento (por si afectó varios lotes)
        total_salida = mov.detalles.aggregate(total=Sum('cantidad'))['total'] or 0

        if total_salida == 0:
            continue

        # Buscamos todas las devoluciones que referencian a este documento
        total_devuelto = Documento.objects.filter(
            documento_referencia=mov,
            tipo_mov='DEVOLUCION'
        ).aggregate(total=Sum('detalles__cantidad'))['total'] or 0

        cantidad_restante = total_salida - total_devuelto
        tiempo_agotado = mov.fecha < limite_devolucion

        # Determinamos el estado
        if cantidad_restante <= 0:
            estado_txt = 'Devolución Completa'
            puede_devolver = False
        elif tiempo_agotado:
            estado_txt = 'Tiempo Expirado'
            puede_devolver = False
        else:
            estado_txt = 'Activo'
            puede_devolver = True

        # Tomamos el nombre del primer detalle para mostrarlo en la lista
        primer_detalle = mov.detalles.first()
        nombre_med = primer_detalle.medicamento.principio_activo if primer_detalle else "Medicamento desconocido"

        historial.append({
            'doc_id': mov.id,
            'medicamento': nombre_med,
            'cantidad_original': total_salida,
            'cantidad_devuelta': total_devuelto,
            'cantidad_restante': cantidad_restante,
            'paciente': mov.id_paciente,
            'fecha': mov.fecha.strftime("%I:%M %p - %d/%b"),
            'puede_devolver': puede_devolver,
            'estado_txt': estado_txt,
            # Enviamos un ID de stock genérico para la animación si es necesario
            'stock_id': primer_detalle.medicamento.id if primer_detalle else None
        })

    return JsonResponse({'historial': historial})


@login_required
def admin_dashboard_view(request):
    """Renderiza el dashboard administrativo según los roles del usuario"""

    roles_disponibles = Group.objects.all()
    if not roles_disponibles.exists():
        for r in ['ADMIN', 'REGENTE', 'ENFERMERA']:
            Group.objects.get_or_create(name=r)
        roles_disponibles = Group.objects.all()
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    es_admin = 'ADMIN' in grupos_usuario
    es_regente = 'REGENTE' in grupos_usuario

    if not (es_admin or es_regente):
        return render(request, 'kardex/error.html',
                      {'mensaje': 'Acceso denegado. Se requieren permisos administrativos.'})

    if es_admin:
        solicitudes = SolicitudStock.objects.select_related('medicamento', 'sede_solicitante').order_by(
            '-fecha_solicitud')
    else:
        solicitudes = SolicitudStock.objects.select_related('medicamento', 'sede_solicitante').filter(
            sede_solicitante=request.user.perfil.ubicacion_asignada).order_by('-fecha_solicitud')

    # Turnos activos: admin ve todas las sedes, regente solo su sede
    if es_admin:
        turnos_activos = TurnoEnfermera.objects.filter(
            activo=True,
            fecha_expiracion__gt=timezone.now()
        ).select_related('enfermera', 'sede').order_by('sede__nombre', '-fecha_inicio')
    else:
        turno = get_turno_activo(request.user.perfil.ubicacion_asignada)
        turnos_activos = [turno] if turno else []

    return render(request, 'kardex/admin_dashboard.html', {
        'es_admin': es_admin,
        'es_regente': es_regente,
        'ubicacion': request.user.perfil.ubicacion_asignada,
        'sedes': Ubicacion.objects.all(),
        'roles': roles_disponibles,
        'solicitudes': solicitudes,
        'usuarios': User.objects.select_related('perfil').prefetch_related('groups').all(),
        'sedes_json': json.dumps(list(Ubicacion.objects.values('id', 'nombre', 'es_bodega_principal'))),
        'medicamentos': Medicamento.objects.filter(
            inventariostock__ubicacion=request.user.perfil.ubicacion_asignada
        ).distinct().order_by('principio_activo'),
        'tipos_medicamento': Medicamento.TIPOS,
        'turno_activo': get_turno_activo(request.user.perfil.ubicacion_asignada),
        'turnos_activos': turnos_activos,
        'conciliaciones': Conciliacion.objects.select_related('carga_rips').order_by('-fecha_conciliacion')[:5],
    })


@login_required
@require_POST
def api_gestion_producto(request):
    """Crea o edita un registro de stock de forma manual (ADMIN y REGENTE)"""
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if not ('ADMIN' in grupos_usuario or 'REGENTE' in grupos_usuario):
        return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)

    try:
        data = json.loads(request.body)
        producto_id = data.get('id')

        # ADMIN puede elegir sede; REGENTE solo su propia sede
        if 'ADMIN' in grupos_usuario:
            ubicacion_id = data.get('ubicacion_id')
            if ubicacion_id:
                ubicacion_actual = Ubicacion.objects.get(id=ubicacion_id)
            else:
                ubicacion_actual = request.user.perfil.ubicacion_asignada
        else:
            ubicacion_actual = request.user.perfil.ubicacion_asignada

        with transaction.atomic():
            tipo_val = data.get('tipo', 'MEDICAMENTO')
            if tipo_val not in ('MEDICAMENTO', 'DISPOSITIVO'):
                tipo_val = 'MEDICAMENTO'

            principio_activo = data.get('principio_activo', '').strip().upper()
            if not principio_activo:
                raise ValueError('El principio activo/nombre es obligatorio.')

            cups_codigo = data.get('cups_codigo', '').strip()
            if not cups_codigo:
                raise ValueError('El código CUPS (RIPS) es obligatorio para todos los productos.')

            forma_farmaceutica = data.get('forma_farmaceutica', '').strip().upper()

            if tipo_val == 'DISPOSITIVO':
                # Para DISPOSITIVO: lookup solo por principio_activo, forma_farmaceutica = 'NO APLICA'
                forma_farmaceutica = 'NO APLICA'
                medicamento, _ = Medicamento.objects.get_or_create(
                    principio_activo=principio_activo,
                    defaults={'forma_farmaceutica': forma_farmaceutica}
                )
                medicamento.forma_farmaceutica = forma_farmaceutica
            else:
                # Para MEDICAMENTO: lookup por principio_activo + forma_farmaceutica + concentracion
                # para que diferentes concentraciones (ej: PROLENE 4-0, 5-0, 3-0) sean registros distintos
                if not forma_farmaceutica:
                    raise ValueError('La forma farmacéutica es obligatoria para medicamentos.')
                lookup_concentracion = data.get('concentracion', '').strip() or None
                medicamento, _ = Medicamento.objects.get_or_create(
                    principio_activo=principio_activo,
                    forma_farmaceutica=forma_farmaceutica,
                    concentracion=lookup_concentracion,
                    defaults={'tipo': tipo_val}
                )

            # Control de nulos para campos únicos
            codigo_ingresado = data.get('codigo', '').strip()
            if codigo_ingresado:
                # Verificar que el código no esté duplicado en otro medicamento
                codigo_qs = Medicamento.objects.filter(codigo=codigo_ingresado)
                if medicamento.pk:
                    codigo_qs = codigo_qs.exclude(pk=medicamento.pk)
                if codigo_qs.exists():
                    raise ValueError(f"El código '{codigo_ingresado}' ya está registrado en otro producto.")
                medicamento.codigo = codigo_ingresado
            else:
                medicamento.codigo = None
            medicamento.tipo = tipo_val
            medicamento.concentracion = data.get('concentracion', medicamento.concentracion) if tipo_val == 'MEDICAMENTO' else None
            medicamento.presentacion = data.get('presentacion', medicamento.presentacion)
            medicamento.laboratorio = data.get('laboratorio', medicamento.laboratorio)
            medicamento.vida_util = data.get('vida_util', medicamento.vida_util)
            medicamento.clasificacion_riesgo = data.get('clasificacion_riesgo', medicamento.clasificacion_riesgo)
            medicamento.cups_codigo = cups_codigo
            medicamento.save()

            lote_raw = data.get('lote')
            lote_ingresado = lote_raw.strip().upper() if lote_raw else ''
            if not lote_ingresado:
                raise ValueError('El número de lote es obligatorio.')

            # Validar que la fecha de vencimiento no sea pasada (si se proporciona)
            from datetime import date, datetime
            fecha_venc = data.get('fecha_vencimiento')
            if fecha_venc:
                if isinstance(fecha_venc, str):
                    try:
                        fecha_venc_date = date.fromisoformat(fecha_venc)
                    except ValueError:
                        fecha_venc_date = datetime.strptime(fecha_venc, '%d/%m/%Y').date()
                else:
                    fecha_venc_date = fecha_venc
                if fecha_venc_date < date.today():
                    raise ValueError("La fecha de vencimiento no puede ser anterior a la fecha actual.")
            else:
                fecha_venc = None

            # Validación de Integridad de Lotes
            if producto_id:
                stock = InventarioStock.objects.get(id=producto_id, ubicacion=ubicacion_actual)
                if stock.lote != lote_ingresado and InventarioStock.objects.filter(ubicacion=ubicacion_actual,
                                                                                   medicamento=medicamento,
                                                                                   lote=lote_ingresado).exists():
                    raise ValueError(f"El lote '{lote_ingresado}' ya pertenece a este medicamento.")
            else:
                if InventarioStock.objects.filter(ubicacion=ubicacion_actual, medicamento=medicamento,
                                                  lote=lote_ingresado).exists():
                    raise ValueError(f"El lote '{lote_ingresado}' ya está registrado.")

                stock = InventarioStock(ubicacion=ubicacion_actual)

            stock.medicamento = medicamento
            stock.lote = lote_ingresado
            stock.fecha_vencimiento = fecha_venc
            stock.cantidad_actual = int(data.get('cantidad'))
            stock.stock_minimo = int(data.get('stock_minimo', 10))
            stock.save()

        return JsonResponse({'status': 'success', 'mensaje': 'Producto guardado correctamente'})

    except ValueError as ve:
        return JsonResponse({'status': 'error', 'mensaje': str(ve)}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


@login_required
@require_POST
def api_carga_masiva(request):
    """Procesa un archivo CSV para cargar inventario de forma masiva (Solo ADMIN)"""

    # 1. Validación de Seguridad Estricta
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos_usuario:
        return JsonResponse({'status': 'error',
                             'mensaje': 'Solo los Administradores Centrales pueden realizar cargas masivas de inventario.'},
                            status=403)

    # 2. Validación de Archivo
    if 'archivo' not in request.FILES:
        return JsonResponse({'status': 'error', 'mensaje': 'No se detectó ningún archivo en la petición.'}, status=400)

    archivo_csv = request.FILES['archivo']
    if not archivo_csv.name.endswith('.csv'):
        return JsonResponse(
            {'status': 'error', 'mensaje': 'Formato inválido. Por favor sube estrictamente un archivo .CSV'},
            status=400)

    try:
        # Importamos el servicio que procesa el Excel (Asegúrate de tener esta función en services.py)
        from .services import procesar_carga_masiva_productos

        # Obtenemos la sede en la que está el administrador
        ubicacion_actual = request.user.perfil.ubicacion_asignada

        # 3. Procesamiento ACID
        total_procesados = procesar_carga_masiva_productos(request.user, archivo_csv)

        return JsonResponse({
            'status': 'success',
            'mensaje': f'¡Carga masiva exitosa! Se procesaron {total_procesados} registros correctamente.'
        })

    except ValueError as ve:
        # Errores específicos (ej: columnas faltantes en el CSV)
        return JsonResponse({'status': 'error', 'mensaje': str(ve)}, status=400)
    except Exception as e:
        # Errores fatales de base de datos
        return JsonResponse({'status': 'error', 'mensaje': f'Error interno procesando el archivo: {str(e)}'},
                            status=400)


@login_required
def descargar_plantilla_carga(request):
    """Descarga la plantilla de carga masiva en formato XLSX"""
    wb = generar_plantilla_xlsx()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_carga_masiva_kardex.xlsx'
    wb.save(response)
    return response


@login_required
def descargar_plantilla_usuarios(request):
    """Descarga plantilla para importar usuarios (XLSX o CSV)"""
    from openpyxl import Workbook

    formato = request.GET.get('formato', 'xlsx').lower()
    filename = f'plantilla_importar_usuarios.{formato}'

    if formato == 'csv':
        import csv
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['nombre', 'documento', 'rol', 'sede'])
        writer.writerow(['Ejemplo Enfermera', '1234567890', 'ENFERMERA', 'Puerto Tejada'])
        writer.writerow(['Ejemplo Regente', '987654321', 'REGENTE', 'FarmaciaSede1'])
        content = output.getvalue()
        response = HttpResponse(content, content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    # XLSX por defecto
    wb = Workbook()
    ws = wb.active
    ws.title = 'Usuarios'
    ws.append(['nombre', 'documento', 'rol', 'sede'])
    ws.append(['Ejemplo Enfermera', '1234567890', 'ENFERMERA', 'Puerto Tejada'])
    ws.append(['Ejemplo Regente', '987654321', 'REGENTE', 'FarmaciaSede1'])
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
@require_POST
def api_carga_masiva_usuarios(request):
    """Importa usuarios desde un archivo Excel o CSV.
    Columnas esperadas: nombre, documento, [rol]
    Roles válidos: ENFERMERA (default), REGENTE, ADMIN
    Clave temporal = documento completo, must_change_password=True
    """
    from django.contrib.auth.models import Group
    from django.contrib.auth.hashers import make_password
    from kardex.models import Ubicacion

    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos_usuario:
        return JsonResponse({'status': 'error', 'mensaje': 'Solo administradores pueden importar usuarios'},
                            status=403)

    if 'archivo' not in request.FILES:
        return JsonResponse({'status': 'error', 'mensaje': 'Debes seleccionar un archivo.'}, status=400)

    archivo = request.FILES['archivo']
    nombre_archivo = archivo.name.lower()

    # Asegurar que existen los grupos
    ROLES_VALIDOS = {'ENFERMERA', 'REGENTE', 'ADMIN'}
    for r in ROLES_VALIDOS:
        Group.objects.get_or_create(name=r)

    # Cache de sedes para evitar consultas repetidas
    from django.core.cache import cache
    _cache_sedes = {}

    def _resolver_sede(nombre_sede):
        if not nombre_sede:
            return None
        key = nombre_sede.strip().lower()
        if key not in _cache_sedes:
            _cache_sedes[key] = Ubicacion.objects.filter(nombre__iexact=nombre_sede.strip()).first()
        return _cache_sedes[key]

    try:
        registros = []

        if nombre_archivo.endswith('.csv'):
            import csv
            import io
            decoded = archivo.read().decode('utf-8-sig')
            # Buscar fila de encabezado real (empieza con 'nombre')
            lines = decoded.split('\n')
            header_idx = None
            for i, line in enumerate(lines):
                if line.strip().lower().startswith('nombre'):
                    header_idx = i
                    break
            if header_idx is not None:
                decoded = '\n'.join(lines[header_idx:])
            reader = csv.DictReader(io.StringIO(decoded))
            for row in reader:
                nombre = (row.get('nombre') or '').strip()
                documento = (row.get('documento') or '').strip()
                documento = ''.join(c for c in documento if c.isdigit())
                rol = (row.get('rol') or '').strip().upper()
                if not rol or rol not in ROLES_VALIDOS:
                    rol = 'ENFERMERA'
                sede_nombre = (row.get('sede') or '').strip()
                if nombre and documento:
                    registros.append({
                        'nombre': nombre,
                        'documento': documento,
                        'rol': rol,
                        'sede': sede_nombre,
                    })

        elif nombre_archivo.endswith('.xlsx'):
            import openpyxl
            import io
            wb = openpyxl.load_workbook(io.BytesIO(archivo.read()))
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row or len(row) < 2:
                    continue
                nombre = str(row[0]).strip() if row[0] else ''
                documento = str(row[1]).strip() if row[1] else ''
                documento = ''.join(c for c in documento if c.isdigit())
                rol = str(row[2]).strip().upper() if len(row) > 2 and row[2] else 'ENFERMERA'
                if rol not in ROLES_VALIDOS:
                    rol = 'ENFERMERA'
                sede_nombre = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                if nombre and documento:
                    registros.append({
                        'nombre': nombre,
                        'documento': documento,
                        'rol': rol,
                        'sede': sede_nombre,
                    })

        else:
            return JsonResponse({'status': 'error', 'mensaje': 'Formato no soportado. Usa .csv o .xlsx'}, status=400)

        if not registros:
            return JsonResponse({'status': 'error', 'mensaje': 'No se encontraron datos válidos en el archivo.'}, status=400)

        creados = 0
        actualizados = 0
        errores = []

        for item in registros:
            nombre, documento, rol = item['nombre'], item['documento'], item['rol']
            sede_nombre = item.get('sede', '')

            # Resolver sede: primero el nombre indicado, fallback a Puerto Tejada, fallback a la primera
            sede_asignada = _resolver_sede(sede_nombre)
            if not sede_asignada:
                sede_asignada = Ubicacion.objects.filter(nombre__icontains='Puerto Tejada').first()
            if not sede_asignada:
                sede_asignada = Ubicacion.objects.order_by('id').first()

            try:
                from django.contrib.auth.models import User
                user, created = User.objects.get_or_create(
                    username=documento,
                    defaults={
                        'first_name': nombre,
                        'last_name': '',
                        'email': '',
                        'password': make_password(documento),
                    }
                )
                if created:
                    PerfilUsuario.objects.create(
                        usuario=user,
                        ubicacion_asignada=sede_asignada,
                        numero_identificacion=documento,
                        must_change_password=True,
                    )
                    user.groups.add(Group.objects.get(name=rol))
                    if rol != 'ENFERMERA':
                        user.groups.add(Group.objects.get(name='ENFERMERA'))
                    creados += 1
                else:
                    user.first_name = nombre
                    user.set_password(documento)
                    user.save()
                    perfil, _ = PerfilUsuario.objects.get_or_create(
                        usuario=user,
                        defaults={
                            'ubicacion_asignada': sede_asignada,
                            'numero_identificacion': documento,
                            'must_change_password': True,
                        }
                    )
                    if not perfil.must_change_password:
                        perfil.must_change_password = True
                        perfil.save()
                    user.groups.clear()
                    user.groups.add(Group.objects.get(name=rol))
                    if rol != 'ENFERMERA':
                        user.groups.add(Group.objects.get(name='ENFERMERA'))
                    actualizados += 1
            except Exception as e:
                errores.append(f'{documento}: {e}')

        mensaje = f'✅ {creados} creados, {actualizados} actualizados.'
        if errores:
            mensaje += f'\n⚠️ {len(errores)} errores: {" | ".join(errores[:5])}'

        return JsonResponse({'status': 'success', 'mensaje': mensaje})

    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': f'Error procesando archivo: {str(e)}'}, status=400)


@login_required
@require_POST
def api_gestion_usuario(request):
    """Crea o edita un usuario y sus roles asignados (Solo ADMIN)"""
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos_usuario:
        return JsonResponse({'status': 'error', 'mensaje': 'Solo administradores pueden gestionar usuarios'},
                            status=403)

    try:
        data = json.loads(request.body)
        user_id = data.get('id')
        roles_seleccionados = data.get('roles', [])
        identificacion = data.get('identificacion', '').strip()

        # Validar unicidad de identificación
        if identificacion:
            duplicado = PerfilUsuario.objects.filter(
                numero_identificacion=identificacion
            )
            if user_id:
                duplicado = duplicado.exclude(usuario_id=user_id)
            if duplicado.exists():
                return JsonResponse({
                    'status': 'error',
                    'mensaje': f'Ya existe un usuario con la identificación "{identificacion}".'
                }, status=400)

        with transaction.atomic():
            if user_id:
                user = User.objects.get(id=user_id)
                user.first_name = data.get('first_name')
                user.last_name = data.get('last_name')
                # Bug 10: Solo actualizar email si se envió un valor no vacío
                if data.get('email'):
                    user.email = data.get('email')
                raw_password = data.get('password', '')
                if raw_password and raw_password.strip():
                    user.set_password(raw_password)
                user.save()

                perfil = user.perfil
                perfil.ubicacion_asignada_id = data.get('ubicacion_id')
                perfil.numero_identificacion = data.get('identificacion')
                perfil.save()
            else:
                user = User.objects.create_user(
                    username=data.get('username'),
                    password=data.get('password') or 'changeme123',
                    first_name=data.get('first_name'),
                    last_name=data.get('last_name'),
                    email=data.get('email', '')
                )
                PerfilUsuario.objects.create(
                    usuario=user,
                    ubicacion_asignada_id=data.get('ubicacion_id'),
                    numero_identificacion=data.get('identificacion')
                )

            # Asignación múltiple de roles (Grupos)
            user.groups.clear()
            for rol_name in roles_seleccionados:
                grupo, _ = Group.objects.get_or_create(name=rol_name)
                user.groups.add(grupo)

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)

@login_required
@require_POST
def api_crear_solicitud(request):
    """Recibe una lista de items (carrito) y crea solicitudes agrupadas por grupo_id"""
    try:
        data = json.loads(request.body)
        items = data.get('items', [])  # [{medicamento_id: X, cantidad: Y}, ...]

        if not items:
            return JsonResponse({'status': 'error', 'mensaje': 'Debe agregar al menos un medicamento.'}, status=400)

        grupo_id = str(uuid.uuid4())
        for item in items:
            SolicitudStock.objects.create(
                medicamento_id=item['medicamento_id'],
                sede_solicitante=request.user.perfil.ubicacion_asignada,
                usuario_solicitante=request.user,
                cantidad_pedida=item.get('cantidad', 50),
                estado='PENDIENTE',
                grupo_id=grupo_id
            )

        return JsonResponse({'status': 'success', 'grupo_id': grupo_id})
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


@login_required
@require_POST
def api_atender_solicitud(request):
    """Despacha TODAS las solicitudes pendientes de un grupo_id"""
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos_usuario:
        return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)
    try:
        data = json.loads(request.body)
        grupo_id = data.get('grupo_id')
        solicitudes = SolicitudStock.objects.filter(grupo_id=grupo_id, estado='PENDIENTE')
        if not solicitudes.exists():
            return JsonResponse({'status': 'error', 'mensaje': 'No hay solicitudes pendientes en este grupo.'}, status=400)

        with transaction.atomic():
            for solicitud in solicitudes.select_related('medicamento', 'sede_solicitante').select_for_update():
                bodega_central = Ubicacion.objects.filter(es_bodega_principal=True).first()
                if bodega_central:
                    total_stock_origen = InventarioStock.objects.filter(
                        ubicacion=bodega_central, medicamento=solicitud.medicamento
                    ).aggregate(total=Sum('cantidad_actual'))['total'] or 0
                    if total_stock_origen < solicitud.cantidad_pedida:
                        raise ValueError(f"Stock insuficiente para {solicitud.medicamento.principio_activo}. Disponible: {total_stock_origen}, Solicitado: {solicitud.cantidad_pedida}")

                    cantidad_restante = solicitud.cantidad_pedida
                    stocks_origen = InventarioStock.objects.filter(
                        ubicacion=bodega_central, medicamento=solicitud.medicamento, cantidad_actual__gt=0
                    ).select_for_update().order_by(Coalesce('fecha_vencimiento', V('9999-12-31')))
                    for s in stocks_origen:
                        if cantidad_restante <= 0: break
                        a_descontar = min(s.cantidad_actual, cantidad_restante)
                        s.cantidad_actual -= a_descontar
                        s.save()
                        cantidad_restante -= a_descontar

                # Sumar a sede destino
                stock = InventarioStock.objects.filter(
                    medicamento=solicitud.medicamento, ubicacion=solicitud.sede_solicitante
                ).order_by(Coalesce('fecha_vencimiento', V('9999-12-31')).desc()).first()
                if stock:
                    stock.cantidad_actual += solicitud.cantidad_pedida
                    stock.save()
                    lote_destino = stock.lote
                else:
                    lote_destino = 'ASIGNADO-CENTRAL'
                    InventarioStock.objects.create(
                        ubicacion=solicitud.sede_solicitante, medicamento=solicitud.medicamento,
                        lote=lote_destino, fecha_vencimiento=timezone.now().date() + timedelta(days=365),
                        cantidad_actual=solicitud.cantidad_pedida, stock_minimo=10
                    )

                doc = Documento.objects.create(
                    tipo_mov='ENTRADA', usuario=request.user, destino=solicitud.sede_solicitante,
                    origen=bodega_central, id_paciente=f"SOL-{solicitud.id}"
                )
                DocumentoDetalle.objects.create(
                    documento=doc, medicamento=solicitud.medicamento, lote=lote_destino, cantidad=solicitud.cantidad_pedida
                )
                solicitud.estado = 'SOLICITADO'
                solicitud.save()

        return JsonResponse({'status': 'success', 'mensaje': f'{solicitudes.count()} items despachados correctamente.'})
    except ValueError as ve:
        return JsonResponse({'status': 'error', 'mensaje': str(ve)}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


# ==============================================================================
# CRUD Ubicacion
# ==============================================================================
@login_required
@require_POST
def api_gestion_ubicacion(request):
    """Crea, edita o elimina una ubicación (solo ADMIN)"""
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos_usuario:
        return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)

    try:
        data = json.loads(request.body)
        accion = data.get('accion', 'guardar')
        ubicacion_id = data.get('id')

        if accion == 'eliminar':
            ubi = Ubicacion.objects.get(id=ubicacion_id)
            # Verificar que no tenga stock ni usuarios asociados
            if InventarioStock.objects.filter(ubicacion=ubi).exists():
                return JsonResponse({'status': 'error', 'mensaje': 'No se puede eliminar: la ubicación tiene inventario asociado.'}, status=400)
            if PerfilUsuario.objects.filter(ubicacion_asignada=ubi).exists():
                return JsonResponse({'status': 'error', 'mensaje': 'No se puede eliminar: la ubicación tiene personal asignado.'}, status=400)
            if SolicitudStock.objects.filter(sede_solicitante=ubi).exists():
                return JsonResponse({'status': 'error', 'mensaje': 'No se puede eliminar: la ubicación tiene solicitudes asociadas.'}, status=400)
            ubi.delete()
            return JsonResponse({'status': 'success', 'mensaje': 'Ubicación eliminada correctamente.'})

        nombre = data.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'status': 'error', 'mensaje': 'El nombre es obligatorio.'}, status=400)

        es_principal = data.get('es_bodega_principal', False)

        if ubicacion_id:
            ubi = Ubicacion.objects.get(id=ubicacion_id)
            ubi.nombre = nombre
            if es_principal:
                # Solo una bodega principal
                Ubicacion.objects.filter(es_bodega_principal=True).exclude(id=ubi.id).update(es_bodega_principal=False)
            ubi.es_bodega_principal = es_principal
            ubi.save()
        else:
            if es_principal:
                Ubicacion.objects.filter(es_bodega_principal=True).update(es_bodega_principal=False)
            Ubicacion.objects.create(nombre=nombre, es_bodega_principal=es_principal)

        return JsonResponse({'status': 'success'})

    except Ubicacion.DoesNotExist:
        return JsonResponse({'status': 'error', 'mensaje': 'Ubicación no encontrada.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


# ==============================================================================
# CRUD ConfiguracionSistema
# ==============================================================================
@login_required
@require_POST
def api_gestion_configuracion(request):
    """Obtiene o actualiza la configuración del sistema (solo ADMIN)"""
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos_usuario:
        return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)

    try:
        data = json.loads(request.body)
        accion = data.get('accion', 'guardar')

        if accion == 'obtener':
            config = ConfiguracionSistema.objects.first()
            if not config:
                config = ConfiguracionSistema.objects.create(horas_limite_devolucion=2)
            return JsonResponse({
                'status': 'success',
                'config': {
                    'id': config.id,
                    'horas_limite_devolucion': config.horas_limite_devolucion,
                    'alertas_habilitadas': config.alertas_habilitadas
                }
            })

        # Guardar
        config = ConfiguracionSistema.objects.first()
        if not config:
            config = ConfiguracionSistema(horas_limite_devolucion=2, alertas_habilitadas=True)

        horas = int(data.get('horas_limite_devolucion', 2))
        if horas < 1:
            return JsonResponse({'status': 'error', 'mensaje': 'Las horas deben ser al menos 1.'}, status=400)
        config.horas_limite_devolucion = horas

        # Toggle de alertas
        if 'alertas_habilitadas' in data:
            config.alertas_habilitadas = bool(data['alertas_habilitadas'])

        config.save()

        return JsonResponse({'status': 'success', 'mensaje': 'Configuración actualizada.'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


# ==============================================================================
@login_required
def api_estado_alertas(request):
    """Endpoint rápido para saber si las alertas están habilitadas"""
    config = ConfiguracionSistema.objects.first()
    habilitadas = config.alertas_habilitadas if config else True
    return JsonResponse({'alertas_habilitadas': habilitadas})


# ==============================================================================
# CRUD Documentos (visor de movimientos)
# ==============================================================================
@login_required
def api_listar_movimientos(request):
    """Lista todos los movimientos del sistema (admin) o de la sede (regente)"""
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    es_admin = 'ADMIN' in grupos_usuario
    es_regente = 'REGENTE' in grupos_usuario

    if not (es_admin or es_regente):
        return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)

    try:
        movimientos = Documento.objects.select_related('usuario', 'origen', 'destino')\
            .prefetch_related('detalles__medicamento').order_by('-fecha')

        if es_regente:
            # Filtrar por sede del regente
            ubi = request.user.perfil.ubicacion_asignada
            movimientos = movimientos.filter(
                models_Q(origen=ubi) | models_Q(destino=ubi)
            )

        data = []
        for mov in movimientos:
            detalles = []
            for det in mov.detalles.all():
                detalles.append({
                    'medicamento': str(det.medicamento),
                    'lote': det.lote,
                    'cantidad': det.cantidad
                })
            data.append({
                'id': mov.id,
                'tipo_mov': mov.tipo_mov,
                'fecha': mov.fecha.strftime('%Y-%m-%d %H:%M'),
                'usuario': mov.usuario.get_full_name() or mov.usuario.username,
                'origen': str(mov.origen) if mov.origen else '-',
                'destino': str(mov.destino) if mov.destino else '-',
                'id_paciente': mov.id_paciente or '-',
                'detalles': detalles
            })

        return JsonResponse({'status': 'success', 'movimientos': data})

    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


# ==============================================================================
# Cancelar SolicitudStock
# ==============================================================================
@login_required
@require_POST
def api_cancelar_solicitud(request):
    """Cancela/elimina TODAS las solicitudes pendientes de un grupo_id"""
    try:
        data = json.loads(request.body)
        grupo_id = data.get('grupo_id')
        solicitudes = SolicitudStock.objects.filter(grupo_id=grupo_id, estado='PENDIENTE')
        if not solicitudes.exists():
            return JsonResponse({'status': 'error', 'mensaje': 'No hay solicitudes pendientes en este grupo.'}, status=400)

        grupos_usuario = request.user.groups.values_list('name', flat=True)
        if not ('ADMIN' in grupos_usuario or solicitudes.first().usuario_solicitante == request.user):
            return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)

        count = solicitudes.count()
        solicitudes.delete()
        return JsonResponse({'status': 'success', 'mensaje': f'{count} solicitud(es) cancelada(s).'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


# ==============================================================================
# Eliminar InventarioStock / Medicamento
# ==============================================================================
@login_required
@require_POST
# ==============================================================================
# TRASLADO MANUAL: Bodega Central → Sede
# ==============================================================================
@login_required
@require_POST
def api_realizar_traslado(request):
    """
    Traslada stock desde la bodega principal a una sede destino.
    Recibe: {medicamento_id, lote, cantidad, sede_destino_id}
    """
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos_usuario:
        return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)

    try:
        data = json.loads(request.body)
        med_id = data.get('medicamento_id')
        lote = data.get('lote', '').strip().upper()
        cantidad = int(data.get('cantidad', 0))
        sede_destino_id = data.get('sede_destino_id')

        if not med_id or not lote or cantidad <= 0 or not sede_destino_id:
            return JsonResponse({'status': 'error', 'mensaje': 'Faltan datos: medicamento, lote, cantidad y sede destino son obligatorios.'}, status=400)

        bodega_principal = Ubicacion.objects.filter(es_bodega_principal=True).first()
        if not bodega_principal:
            return JsonResponse({'status': 'error', 'mensaje': 'No hay una bodega principal configurada.'}, status=400)

        sede_destino = Ubicacion.objects.get(id=sede_destino_id)
        if sede_destino.id == bodega_principal.id:
            return JsonResponse({'status': 'error', 'mensaje': 'No se puede trasladar a la misma bodega principal.'}, status=400)

        medicamento = Medicamento.objects.get(id=med_id)

        with transaction.atomic():
            stock_origen = InventarioStock.objects.select_for_update().get(
                ubicacion=bodega_principal,
                medicamento=medicamento,
                lote=lote
            )

            if stock_origen.cantidad_actual < cantidad:
                return JsonResponse({
                    'status': 'error',
                    'mensaje': f'Stock insuficiente en bodega central. Disponible: {stock_origen.cantidad_actual}, Solicitado: {cantidad}.'
                }, status=400)

            stock_destino, _ = InventarioStock.objects.get_or_create(
                ubicacion=sede_destino,
                medicamento=medicamento,
                lote=lote,
                defaults={
                    'cantidad_actual': 0,
                    'fecha_vencimiento': stock_origen.fecha_vencimiento,
                    'stock_minimo': stock_origen.stock_minimo or 10,
                }
            )

            stock_origen.cantidad_actual -= cantidad
            stock_origen.save()

            stock_destino.cantidad_actual += cantidad
            stock_destino.save()

            doc = Documento.objects.create(
                tipo_mov='TRASLADO',
                origen=bodega_principal,
                destino=sede_destino,
                usuario=request.user,
            )
            DocumentoDetalle.objects.create(
                documento=doc,
                medicamento=medicamento,
                lote=lote,
                cantidad=cantidad,
            )

        return JsonResponse({
            'status': 'success',
            'mensaje': f'✅ {cantidad} unidades de {medicamento.principio_activo} trasladadas a {sede_destino.nombre}.'
        })

    except Ubicacion.DoesNotExist:
        return JsonResponse({'status': 'error', 'mensaje': 'Sede destino no encontrada.'}, status=404)
    except Medicamento.DoesNotExist:
        return JsonResponse({'status': 'error', 'mensaje': 'Medicamento no encontrado.'}, status=404)
    except InventarioStock.DoesNotExist:
        return JsonResponse({'status': 'error', 'mensaje': f'No hay stock del lote "{lote}" en la bodega principal.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


def api_eliminar_stock(request):
    """Elimina un registro de stock, y opcionalmente el medicamento si queda huérfano"""
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos_usuario:
        return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)

    try:
        data = json.loads(request.body)
        stock_id = data.get('stock_id')

        stock = InventarioStock.objects.select_related('medicamento').get(id=stock_id)
        medicamento = stock.medicamento

        if stock.cantidad_actual > 0:
            return JsonResponse({'status': 'error', 'mensaje': f'No se puede eliminar: el lote tiene {stock.cantidad_actual} unidades. Debe agotar el stock primero o ajustar a 0.'}, status=400)

        stock.delete()

        # Si el medicamento ya no tiene stock en ninguna ubicación, lo eliminamos
        if not InventarioStock.objects.filter(medicamento=medicamento).exists():
            # Verificar que no tenga movimientos asociados
            if not DocumentoDetalle.objects.filter(medicamento=medicamento).exists():
                medicamento.delete()

        return JsonResponse({'status': 'success', 'mensaje': 'Registro de stock eliminado.'})

    except InventarioStock.DoesNotExist:
        return JsonResponse({'status': 'error', 'mensaje': 'Stock no encontrado.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


# ==============================================================================
# Eliminar Usuario
# ==============================================================================
@login_required
@require_POST
def api_eliminar_usuario(request):
    """Elimina un usuario del sistema (solo ADMIN, no a sí mismo)"""
    grupos_usuario = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos_usuario:
        return JsonResponse({'status': 'error', 'mensaje': 'No autorizado'}, status=403)

    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')

        if int(user_id) == request.user.id:
            return JsonResponse({'status': 'error', 'mensaje': 'No puedes eliminarte a ti mismo.'}, status=400)

        user = User.objects.get(id=user_id)

        # Verificar que no tenga movimientos asociados
        if Documento.objects.filter(usuario=user).exists():
            return JsonResponse({'status': 'error', 'mensaje': 'No se puede eliminar: el usuario tiene movimientos registrados. Desactívelo en su lugar.'}, status=400)

        # Eliminar perfil y usuario
        PerfilUsuario.objects.filter(usuario=user).delete()
        user.delete()

        return JsonResponse({'status': 'success', 'mensaje': 'Usuario eliminado correctamente.'})

    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'mensaje': 'Usuario no encontrado.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)


# ==============================================================================
# API: CARGA RIPS (desde la web)
# ==============================================================================


@login_required
@require_POST
def api_cargar_rips(request):
    """Sube y procesa el reporte201 RIPS desde la web"""
    grupos = request.user.groups.values_list('name', flat=True)
    if 'ADMIN' not in grupos:
        return JsonResponse({'status': 'error', 'mensaje': 'Solo ADMIN puede cargar RIPS'}, status=403)

    if 'archivo_rips' not in request.FILES:
        return JsonResponse({'status': 'error', 'mensaje': 'No se envió ningún archivo'}, status=400)

    archivo = request.FILES['archivo_rips']
    if not archivo.name.endswith('.csv'):
        return JsonResponse({'status': 'error', 'mensaje': 'El archivo debe ser .csv'}, status=400)

    try:
        from .services import procesar_importacion_rips, ejecutar_conciliacion

        carga, mensaje, success = procesar_importacion_rips(archivo)
        if not success:
            return JsonResponse({'status': 'error', 'mensaje': mensaje})

        conciliacion = ejecutar_conciliacion(carga)

        return JsonResponse({
            'status': 'success',
            'mensaje': mensaje,
            'conciliacion_id': conciliacion.id,
            'data': {
                'coincidencias': conciliacion.coincidencias,
                'no_facturados': conciliacion.no_facturados,
                'no_despachados': conciliacion.no_despachados,
                'total_kardex': conciliacion.total_salidas_kardex,
                'total_rips': conciliacion.total_registros_rips,
            }
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)


# ==============================================================================
# VISTAS DE CONCILIACIÓN RIPS
# ==============================================================================


@login_required
def conciliacion_lista(request):
    """Listado de conciliaciones realizadas"""
    from .models import Conciliacion, CargaRIPS

    conciliaciones = Conciliacion.objects.all().select_related('carga_rips')[:20]
    cargas_sin_conciliar = CargaRIPS.objects.filter(estado='CARGADA')

    return render(request, 'kardex/conciliacion_lista.html', {
        'conciliaciones': conciliaciones,
        'cargas_sin_conciliar': cargas_sin_conciliar,
    })


@login_required
def conciliacion_detalle(request, conciliacion_id):
    """Detalle de una conciliación"""
    from .models import Conciliacion, DetalleConciliacion

    conciliacion = Conciliacion.objects.get(id=conciliacion_id)
    estado_filtro = request.GET.get('estado', '')

    detalles = DetalleConciliacion.objects.filter(conciliacion=conciliacion)
    if estado_filtro:
        detalles = detalles.filter(estado=estado_filtro)
    detalles = detalles.order_by('-fecha')[:100]

    return render(request, 'kardex/conciliacion_detalle.html', {
        'conciliacion': conciliacion,
        'detalles': detalles,
        'estado_filtro': estado_filtro,
    })


@login_required
def conciliacion_exportar_excel(request, conciliacion_id):
    """Exporta las discrepancias de una conciliación a Excel"""
    from .models import Conciliacion, DetalleConciliacion
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    conciliacion = Conciliacion.objects.get(id=conciliacion_id)
    detalles = DetalleConciliacion.objects.filter(conciliacion=conciliacion)

    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliación"

    fuente_titulo = Font(bold=True, size=14, name='Arial')
    fuente_header = Font(bold=True, size=10, name='Arial', color='FFFFFF')
    fuente_normal = Font(size=10, name='Arial')
    alineacion_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alineacion_izq = Alignment(horizontal='left', vertical='center', wrap_text=True)
    borde_fino = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fondo_rojo = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    fondo_verde = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    fondo_amarillo = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    fondo_header = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')

    ws.merge_cells('A1:H1')
    ws['A1'] = f'CONCILIACIÓN KARDEX vs RIPS - {conciliacion.periodo_inicio} a {conciliacion.periodo_fin}'
    ws['A1'].font = fuente_titulo
    ws['A1'].alignment = Alignment(horizontal='center')

    row = 3
    ws.cell(row=row, column=1, value='RESUMEN').font = Font(bold=True, size=12)
    row = 4
    ws.cell(row=row, column=1, value='Total Kardex:'); ws.cell(row=row, column=2, value=conciliacion.total_salidas_kardex)
    row = 5
    ws.cell(row=row, column=1, value='Total RIPS:'); ws.cell(row=row, column=2, value=conciliacion.total_registros_rips)
    row = 6
    ws.cell(row=row, column=1, value='Coincidencias:'); ws.cell(row=row, column=2, value=conciliacion.coincidencias)
    row = 7
    ws.cell(row=row, column=1, value='No Facturados:'); ws.cell(row=row, column=2, value=conciliacion.no_facturados)
    row = 8
    ws.cell(row=row, column=1, value='No Despachados:'); ws.cell(row=row, column=2, value=conciliacion.no_despachados)

    row_start = 10
    headers = ['Estado', 'Medicamento', 'Paciente ID', 'Paciente',
               'Cant. Kardex', 'Cant. RIPS', 'Fecha', 'Observacion']
    for col, h in enumerate(headers, 1):
        celda = ws.cell(row=row_start, column=col, value=h)
        celda.font = fuente_header
        celda.alignment = alineacion_centro
        celda.fill = fondo_header
        celda.border = borde_fino

    for i, det in enumerate(detalles, start=row_start + 1):
        colores = {
            'COINCIDE': fondo_verde,
            'NO_FACTURADO': fondo_rojo,
            'NO_DESPACHADO': fondo_amarillo,
            'CANTIDAD_DIF': fondo_amarillo,
        }
        ws.cell(row=i, column=1, value=dict(DetalleConciliacion.ESTADOS).get(det.estado, det.estado))
        ws.cell(row=i, column=1).fill = colores.get(det.estado, fondo_verde)
        ws.cell(row=i, column=2, value=det.medicamento_nombre)
        ws.cell(row=i, column=3, value=det.paciente_identificacion)
        ws.cell(row=i, column=4, value=det.paciente_nombre)
        ws.cell(row=i, column=5, value=det.cantidad_kardex)
        ws.cell(row=i, column=6, value=det.cantidad_rips)
        ws.cell(row=i, column=7, value=det.fecha.strftime('%Y-%m-%d %H:%M') if det.fecha else '')
        ws.cell(row=i, column=8, value=det.observacion)

        for col in range(1, 9):
            celda = ws.cell(row=i, column=col)
            celda.font = fuente_normal
            celda.alignment = alineacion_centro if col in (1, 3, 5, 6, 7) else alineacion_izq
            celda.border = borde_fino

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 50

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=conciliacion_{conciliacion.id}.xlsx'
    wb.save(response)
    return response